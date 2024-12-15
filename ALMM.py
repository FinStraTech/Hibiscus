import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime 

class ALMM :
    def __init__(self, data_import: pd.DataFrame, ref_entite_path: str, ref_transfo_path: str, ref_almm_path: str, ref_adf_almm_path: str, ref_dzone_almm_path:str, run_timestamp: str, export_type : str):

        self.data = data_import

        # Charger et prétraiter les fichiers de référence
        self.ref_entite = self.preprocess_ref_entite(ref_entite_path)
        self.ref_transfo = self.preprocess_ref_transfo(ref_transfo_path)
        self.ref_almm = self.preprocess_ref_almm(ref_almm_path)
        self.ref_adf_almm = self.preprocess_ref_adf_almm(ref_adf_almm_path)
        self.ref_dzone_almm = self.preprocess_ref_dzone_almm(ref_dzone_almm_path)
        self.run_timestamp = run_timestamp
        self.export_type = export_type

    def _save_import_files(self, filtered_data, import_folder, export_type):
        """
        Sauvegarde les fichiers d'import dans le dossier spécifié par devise (ALL, EUR, USD) 
        et vérifie que les fichiers générés ne sont pas corrompus.

        :param filtered_data: DataFrame filtré.
        :param import_folder: Dossier où sauvegarder les fichiers.
        :param export_type: Type d'export (ALL, BILAN, CONSO).
        :return: Dictionnaire contenant les chemins des fichiers générés.
        """
        saved_files = {}

        for currency in ["ALL", "EUR", "USD"]:
            if currency == "ALL":
                data_to_save = filtered_data
            else:
                data_to_save = filtered_data[filtered_data["D_CU"] == currency]

            # Vérifications avant sauvegarde
            if data_to_save.empty:
                print(f"Aucune donnée trouvée pour la devise {currency} dans {export_type}.")
                continue

            file_name = f"VIEW_{export_type}_IG_{currency}.xlsx"
            file_path = os.path.join(import_folder, file_name)

            try:
                # Étape 1 : Sauvegarder le fichier Excel
                data_to_save.to_excel(file_path, index=False, engine="xlsxwriter")
                print(f"Fichier généré : {file_path}")

                # Étape 2 : Valider que le fichier peut être relu correctement
                try:
                    test_read = pd.read_excel(file_path, engine="openpyxl")
                    if test_read.empty and not data_to_save.empty:
                        raise ValueError(f"Le fichier {file_path} est corrompu (lecture vide après écriture).")
                    if not data_to_save.equals(test_read):
                        raise ValueError(f"Le fichier {file_path} est corrompu (données lues non identiques à celles écrites).")
                except Exception as e:
                    raise ValueError(f"Validation échouée pour le fichier {file_path}: {e}")

                # Ajouter le fichier validé à la liste des fichiers sauvegardés
                saved_files[currency] = file_path

            except Exception as e:
                print(f"Erreur lors de la sauvegarde ou de la validation du fichier {file_path}: {e}")
                # Nettoyer le fichier corrompu s'il existe
                if os.path.exists(file_path):
                    os.remove(file_path)
                print(f"Fichier corrompu supprimé : {file_path}")

        return saved_files

    
    def preprocess_data(self, export_type="ALL", currency="ALL", entity="ALL"):
        """
        Nettoie et convertit les types des colonnes dans les données, génère les fichiers d'import
        pour BILAN, CONSO, ALL, et gère les étapes spécifiques pour GRAN.

        :param export_type: Type d'export choisi par l'utilisateur (ALL, BILAN, CONSO, GRAN).
        :param currency: Devise à filtrer (ALL, EUR, USD).
        :param entity: Entité à filtrer ou ALL.
        :return: Chemins des fichiers sauvegardés ou données filtrées pour GRAN.
        """
        # Création du dossier d'import
        import_folder = f"./imports/import_{self.run_timestamp}"
        os.makedirs(import_folder, exist_ok=True)

        # Suppression des lignes totalement vides
        self.data = self.data.dropna(how="all")
        self.data = self.data[~self.data.apply(lambda row: all(row == ""), axis=1)]

        # Définition des types de colonnes
        column_types = {
            "D_CA": "string",
            "D_DP": "float64",
            "D_PE": "float64",
            "D_RU": "string",
            "D_AC": "string",
            "D_FL": "string",
            "D_CU": "string",
            "D_ZONE": "string",
            "P_AMOUNT": "Int64",
            "D_T1": "string"
        }

        # Conversion des types de colonnes
        for col, dtype in column_types.items():
            if col in self.data.columns:
                try:
                    if dtype == "Int64":
                        self.data[col] = pd.to_numeric(self.data[col], errors="coerce").astype("Int64")
                    else:
                        self.data[col] = self.data[col].astype(dtype)
                except Exception as e:
                    print(f"Erreur lors de la conversion de la colonne {col} en {dtype}: {e}")

        # Étape 1 : Filtrage spécifique pour GRAN
        if export_type == "GRAN":
            if currency == "ALL":
                raise ValueError("Pour un export de type GRAN, une devise spécifique doit être fournie.")

            print(f"Filtrage des données pour la devise '{currency}'...")
            filtered_data_currency = self.data[self.data["D_CU"] == currency]

            if filtered_data_currency.empty:
                raise ValueError(f"Aucune donnée trouvée pour la devise '{currency}'.")

            return filtered_data_currency

        # Étape 2 : Génération des fichiers pour BILAN, CONSO, et ALL
        generated_files = {}
        if export_type in ["ALL", "BILAN"]:
            filtered_bilan = self.data[self.data["D_T1"] == "INTER"]
            generated_files.update(self._save_import_files(filtered_bilan, import_folder, "BILAN"))
        if export_type in ["ALL", "CONSO"]:
            filtered_conso = self.data[self.data["D_T1"] != "INTER"]
            generated_files.update(self._save_import_files(filtered_conso, import_folder, "CONSO"))
        if export_type == "ALL":
            generated_files.update(self._save_import_files(self.data, import_folder, "ALL"))

        print(f"Fichiers d'import sauvegardés dans : {import_folder}")
        return generated_files

    def _save_import_files(self, filtered_data, import_folder, export_type):
        """
        Sauvegarde les fichiers d'import dans le dossier spécifié par devise (ALL, EUR, USD).

        :param filtered_data: DataFrame filtré.
        :param import_folder: Dossier où sauvegarder les fichiers.
        :param export_type: Type d'export (ALL, BILAN, CONSO).
        :return: Dictionnaire contenant les chemins des fichiers générés.
        """
        saved_files = {}
        for currency in ["ALL", "EUR", "USD"]:
            if currency == "ALL":
                data_to_save = filtered_data
            else:
                data_to_save = filtered_data[filtered_data["D_CU"] == currency]

            # Vérifications avant sauvegarde
            if data_to_save.empty:
                print(f"Aucune donnée trouvée pour la devise {currency} dans {export_type}.")
                continue

            file_name = f"VIEW_{export_type}_IG_{currency}.xlsx"
            file_path = os.path.join(import_folder, file_name)
            try:
                data_to_save.to_excel(file_path, index=False, engine="xlsxwriter")
                print(f"Fichier généré : {file_path}")
                saved_files[currency] = file_path
            except Exception as e:
                print(f"Erreur lors de la génération du fichier {file_path}: {e}")

        return saved_files


    
    @staticmethod
    def preprocess_ref_entite(file_path: str) -> pd.DataFrame:
        """
        Prétraitement pour Ref_Entite.xlsx :
        - Supprime les lignes ayant une valeur nulle dans la colonne 'd_ru'.
        - Ajoute le préfixe 'Ref_Entite.' à tous les noms de colonnes.
        """
        df = pd.read_excel(file_path)
        df = df.dropna(subset=['d_ru'])  # Supprime les lignes où 'd_ru' est null

        # Renommer les colonnes en ajoutant le préfixe 'Ref_Entite.'
        df = df.rename(columns=lambda col: f"Ref_Entite.{col}")
        return df

    @staticmethod
    def preprocess_ref_transfo(file_path: str) -> pd.DataFrame:
        df = pd.read_excel(file_path)
        df['Transfo_aggregate_L1'] = df['Transfo_aggregate_L1'].astype(str)  # Convertit en texte
        df = df.drop_duplicates(subset=['Transfo_aggregate_L1'])  # Supprime les doublons

        # Renommer les colonnes en ajoutant le préfixe 'Ref_Transfo_L1.'
        df = df.rename(columns=lambda col: f"Ref_Transfo_L1.{col}")
        return df
    
    @staticmethod
    def preprocess_ref_almm(file_path: str) -> pd.DataFrame:

        df = pd.read_excel(file_path)

        # Changer les types des colonnes
        column_types = {
            "Compte Transfo": "string",
            "Ligne_NSFR": "string",
            "PCT_NSFR": "string",
        }

        for col, dtype in column_types.items():
            if col in df.columns:
                try:
                    df[col] = df[col].astype(dtype)
                except Exception as e:
                    print(f"Erreur lors de la conversion de la colonne {col} en {dtype}: {e}")

        # Renommer les colonnes avec le préfixe 'Ref_NSFR.'
        df = df.rename(columns=lambda col: f"Ref_NSFR.{col}")

        return df

    @staticmethod
    def preprocess_ref_adf_almm(file_path: str) -> pd.DataFrame:
        df = pd.read_excel(file_path)

        column_types = {
            "D_ru": "string",
            "Entité": "string",
            "D_ac": "string",
            "Indicator_Ligne": "string",
            "Indicator_ADF": "Int64",
            "Indicator_ADF_0-6M": "float64",
            "Indicator_ADF_6-12M": "float64",
            "Indicator_ADF_>1Y": "float64",
        }

        for col, dtype in column_types.items():
            if col in df.columns:
                try:
                    if dtype == "float64":
                        # Remplacer les valeurs non numériques
                        df[col] = pd.to_numeric(df[col].replace("NOT APPLICABLE", None), errors='coerce')
                    elif dtype == "Int64":
                        df[col] = pd.to_numeric(df[col], errors='coerce').astype("Int64")
                    else:
                        df[col] = df[col].astype(dtype)
                except Exception as e:
                    print(f"Erreur lors de la conversion de la colonne {col} en {dtype}: {e}")

        # Renommer les colonnes avec le préfixe Ref_ADF_NSFR.
        df = df.rename(columns=lambda x: f"Ref_ADF_NSFR.{x}")

        return df

    @staticmethod
    def preprocess_ref_dzone_almm(file_path: str) -> pd.DataFrame:

        # Charger le fichier
        df = pd.read_excel(file_path)

        # Changer les types des colonnes
        column_types = {
            "Colonne1": "string",
            "Colonne2": "string",
        }
        for col, dtype in column_types.items():
            if col in df.columns:
                try:
                    df[col] = df[col].astype(dtype)
                except Exception as e:
                    print(f"Erreur lors de la conversion de la colonne {col} en {dtype}: {e}")

        # Renommer les colonnes
        rename_columns = {
            "Colonne1": "D_ZONE",
            "Colonne2": "NSFR_Bucket",
        }
        df = df.rename(columns=rename_columns)

        # Ignorer la première ligne
        df = df.iloc[1:].reset_index(drop=True)

        # Ajouter le préfixe 'Ref_DZONE_NSFR.' aux colonnes
        df = df.rename(columns=lambda col: f"Ref_DZONE_NSFR.{col}")

        return df

    def filter_and_join_ref_entite(self,preprocessed_data):

        # 2.2. Filtrer les données
        filtered_data = preprocessed_data[
            (preprocessed_data["D_FL"] != "T99") & (preprocessed_data["D_ZONE"].notna())
        ]
        
        # 2.3. Joindre la table principale filtrée avec Ref_Entite
        joined_data = pd.merge(
            filtered_data,  # Table principale filtrée
            self.ref_entite,  # Table secondaire Ref_Entite
            left_on="D_RU",  # Colonne de jointure dans la table principale
            right_on="Ref_Entite.d_ru",  # Colonne de jointure dans la table secondaire
            how="left",  # Jointure externe gauche
        )

        # Retourner les données après jointure
        return joined_data

    def join_with_ref_transfo(self, filtered_data: pd.DataFrame):
    
        # Effectuer la jointure
        joined_data = pd.merge(
            filtered_data,  # Table principale (déjà filtrée et jointe avec Ref_Entite)
            self.ref_transfo,  # Référence Ref_Transfo_L1 (prétraitée dynamiquement)
            left_on="D_AC",  # Colonne de la table principale
            right_on="Ref_Transfo_L1.Transfo_aggregate_L1",  # Colonne de la référence
            how="left",  # Jointure externe gauche
        )

        # Filtrer les lignes où Transfo_aggregate_L1 n'est pas null
        filtered_joined_data = joined_data[joined_data["Ref_Transfo_L1.Transfo_aggregate_L1"].notna()]

        # Retourner les données après jointure et filtrage
        return filtered_joined_data

    
    def join_with_ref_dzone_almm(self, filtered_data: pd.DataFrame) -> pd.DataFrame:

        # Vérifier que les colonnes nécessaires sont présentes
        if "D_ZONE" not in filtered_data.columns:
            raise ValueError("La colonne 'D_ZONE' est manquante dans le DataFrame principal.")
        if "Ref_DZONE_NSFR.D_ZONE" not in self.ref_dzone_almm.columns:
            raise ValueError("La colonne 'Ref_DZONE_NSFR.D_ZONE' est manquante dans la table Ref_DZONE_NSFR.")

        # Effectuer la jointure
        joined_data = pd.merge(
            filtered_data,  # Table principale
            self.ref_dzone_almm,  # Référence Ref_DZONE_NSFR
            left_on="D_ZONE",  # Colonne de la table principale
            right_on="Ref_DZONE_NSFR.D_ZONE",  # Colonne de la référence
            how="left",  # Jointure externe gauche
        )

        return joined_data

    def join_with_ref_almm(self, filtered_data: pd.DataFrame) -> pd.DataFrame:

        # Vérifier que les colonnes nécessaires sont présentes
        if "D_AC" not in filtered_data.columns:
            raise ValueError("La colonne 'D_AC' est manquante dans le DataFrame principal.")
        if "Ref_NSFR.Compte Transfo" not in self.ref_almm.columns:
            raise ValueError("La colonne 'Ref_NSFR.Compte Transfo' est manquante dans la table Ref_NSFR.")

        # Effectuer la jointure
        joined_data = pd.merge(
            filtered_data,  # Table principale
            self.ref_almm,  # Référence Ref_NSFR
            left_on="D_AC",  # Colonne de la table principale
            right_on="Ref_NSFR.Compte Transfo",  # Colonne de la référence
            how="left",  # Jointure externe gauche
        )

        # Filtrer les lignes où Ref_NSFR.Ligne_NSFR n'est pas null
        if "Ref_NSFR.Ligne_NSFR" not in joined_data.columns:
            raise ValueError("La colonne 'Ref_NSFR.Ligne_NSFR' est manquante dans le DataFrame après jointure.")
        filtered_joined_data = joined_data[joined_data["Ref_NSFR.Ligne_NSFR"].notna()]

        return filtered_joined_data
    
    def group_and_sum_unadjusted_p_amount(self, data: pd.DataFrame) -> pd.DataFrame:
    
        # Colonnes utilisées pour le regroupement
        group_columns = [
            "Ref_Entite.entité", 
            "D_AC", 
            "Ref_DZONE_NSFR.NSFR_Bucket", 
            "Ref_NSFR.Ligne_NSFR"
        ]

        # Vérifier que toutes les colonnes nécessaires sont présentes
        required_columns = group_columns + ["P_AMOUNT"]
        for col in required_columns:
            if col not in data.columns:
                raise ValueError(f"La colonne '{col}' est manquante dans le DataFrame.")

        # Regrouper les données et calculer la somme
        grouped_data = (
            data.groupby(group_columns, as_index=False)
            .agg(Unadjusted_P_Amount=("P_AMOUNT", "sum"))
        )

        return grouped_data

    def pivot_and_reorder(self, data: pd.DataFrame) -> pd.DataFrame:
        
        # Vérifier que toutes les colonnes nécessaires sont présentes
        required_columns = [
            "Ref_Entite.entité",
            "D_AC",
            "Ref_NSFR.Ligne_NSFR",
            "Ref_DZONE_NSFR.NSFR_Bucket",
            "Unadjusted_P_Amount",
        ]
        for col in required_columns:
            if col not in data.columns:
                raise ValueError(f"La colonne '{col}' est manquante dans le DataFrame.")

        # Pivoter les données
        pivoted_data = data.pivot_table(
            index=["Ref_Entite.entité", "D_AC", "Ref_NSFR.Ligne_NSFR"],  # Colonnes fixes
            columns="Ref_DZONE_NSFR.NSFR_Bucket",  # Colonne à pivoter
            values="Unadjusted_P_Amount",  # Valeur à agréger
            aggfunc="sum",  # Fonction d'agrégation
            fill_value=0,  # Remplir les valeurs manquantes par 0
        ).reset_index()

        # Réorganiser les colonnes
        desired_order = [
            "Ref_Entite.entité",
            "D_AC",
            "Ref_NSFR.Ligne_NSFR",
            "0-6M",
            "6-12M",
            ">1Y",
        ]
        for col in desired_order:
            if col not in pivoted_data.columns:
                raise ValueError(f"La colonne '{col}' est manquante dans le DataFrame pivoté.")

        reordered_data = pivoted_data[desired_order]

        return reordered_data

    
    def join_with_ref_adf_almm(self, data: pd.DataFrame) -> pd.DataFrame:

        # Vérifier que les colonnes nécessaires sont présentes
        if "D_AC" not in data.columns or "Ref_NSFR.Ligne_NSFR" not in data.columns:
            raise ValueError("Les colonnes 'D_AC' ou 'Ref_NSFR.Ligne_NSFR' sont manquantes dans la table principale.")
        if "Ref_ADF_NSFR.D_ac" not in self.ref_adf_almm.columns or "Ref_ADF_NSFR.Indicator_Ligne" not in self.ref_adf_almm.columns:
            raise ValueError("Les colonnes 'Ref_ADF_NSFR.D_ac' ou 'Ref_ADF_NSFR.Indicator_Ligne' sont manquantes dans la table Ref_ADF_NSFR.")

        # Effectuer la jointure
        joined_data = pd.merge(
            data,  # Table principale
            self.ref_adf_almm,  # Référence Ref_ADF_NSFR
            left_on=["D_AC", "Ref_NSFR.Ligne_NSFR"],  # Colonnes de la table principale
            right_on=["Ref_ADF_NSFR.D_ac", "Ref_ADF_NSFR.Indicator_Ligne"],  # Colonnes de la référence
            how="left",  # Jointure externe gauche
        )

        return joined_data

    def add_adjusted_amounts(self, data: pd.DataFrame) -> pd.DataFrame:
        # Vérifier que toutes les colonnes nécessaires sont présentes
        required_columns = [
            "0-6M", 
            "6-12M", 
            ">1Y", 
            "Ref_ADF_NSFR.Indicator_ADF_0-6M", 
            "Ref_ADF_NSFR.Indicator_ADF_6-12M", 
            "Ref_ADF_NSFR.Indicator_ADF_>1Y"
        ]
        for col in required_columns:
            if col not in data.columns:
                raise ValueError(f"La colonne '{col}' est manquante dans le DataFrame.")

        # Ajouter les colonnes calculées
        data["P_Adjusted_Amount_0-6M"] = data["0-6M"] * data["Ref_ADF_NSFR.Indicator_ADF_0-6M"]
        data["P_Adjusted_Amount_6-12M"] = data["6-12M"] * data["Ref_ADF_NSFR.Indicator_ADF_6-12M"]
        data["P_Adjusted_Amount_>1Y"] = data[">1Y"] * data["Ref_ADF_NSFR.Indicator_ADF_>1Y"]

        # Colonnes à supprimer
        columns_to_drop = [
            "Ref_ADF_NSFR.D_ru",
            "Ref_ADF_NSFR.D_ac",
            "Ref_ADF_NSFR.Indicator_Ligne",
            "Ref_ADF_NSFR.Indicator_ADF",
        ]
        for col in columns_to_drop:
            if col in data.columns:
                data = data.drop(columns=col)

        return data

    
    def save_excel_with_structure(
        self,
        processed_data: dict,  # Clé : devise, Valeur : DataFrame
        excel_file_path: str,
        entity_list: list,
        run_timestamp: str,
        export_type: str,
        base_output_dir: str = "output",
        entity: str = None,  # Spécifique pour GRAN
        currency: str = "ALL"  # Spécifique pour GRAN
    ):
        """
        Sauvegarde les fichiers Excel selon une structure hiérarchique.

        :param processed_data: Données traitées (dict avec clés comme les devises et valeurs comme DataFrames).
        :param excel_file_path: Chemin du fichier Excel de base.
        :param entity_list: Liste des noms d'entités à filtrer (utilisé pour BILAN, CONSO, ALL).
        :param run_timestamp: Timestamp du traitement.
        :param export_type: Type d'export (ALL, BILAN, CONSO, GRAN).
        :param base_output_dir: Répertoire de sortie.
        :param entity: Nom de l'entité (spécifique pour GRAN).
        :param currency: Devise (spécifique pour GRAN).
        """
        base_folder = os.path.join(base_output_dir, f"RUN_{run_timestamp}_{export_type}")
        os.makedirs(base_folder, exist_ok=True)

        # Traitement pour BILAN et CONSO
        if export_type in ["BILAN", "CONSO"]:
            for currency, data in processed_data.items():
                # Vérifier que `data` est bien un DataFrame
                if not isinstance(data, pd.DataFrame):
                    print(f"Les données pour la devise '{currency}' ne sont pas un DataFrame. Traitement ignoré.")
                    continue

                currency_folder = os.path.join(base_folder, f"{export_type}_{currency}")
                os.makedirs(currency_folder, exist_ok=True)

                all_entities_folder = os.path.join(currency_folder, "Reports_all_entities")
                os.makedirs(all_entities_folder, exist_ok=True)

                by_entity_folder = os.path.join(currency_folder, "Reports_by_entity")
                os.makedirs(by_entity_folder, exist_ok=True)

                # Sauvegarder les fichiers globaux
                global_file = os.path.join(all_entities_folder, f"ALMM_{export_type}_{currency}_All_Entities.xlsx")
                self.save_to_excel(data, excel_file_path, global_file)

                # Sauvegarder par entité
                for entity in entity_list:
                    entity_data = data[data["Ref_Entite.entité"] == entity]
                    entity_folder = os.path.join(by_entity_folder, entity)
                    os.makedirs(entity_folder, exist_ok=True)

                    if not entity_data.empty:
                        entity_file = os.path.join(entity_folder, f"ALMM_{export_type}_{currency}_{entity}.xlsx")
                        self.save_to_excel(entity_data, excel_file_path, entity_file)
                        print(f"Fichier sauvegardé : {entity_file}")

        # Traitement pour ALL
        elif export_type == "ALL":
            # Vérifier que `processed_data` est un dictionnaire de DataFrame
            for currency, data in processed_data.items():
                if not isinstance(data, pd.DataFrame):
                    print(f"Les données pour la devise '{currency}' ne sont pas un DataFrame. Traitement ignoré.")
                    continue

                all_entities_folder = os.path.join(base_folder, "Reports_all_entities")
                os.makedirs(all_entities_folder, exist_ok=True)

                by_entity_folder = os.path.join(base_folder, "Reports_by_entity")
                os.makedirs(by_entity_folder, exist_ok=True)

                # Sauvegarder les fichiers globaux
                global_file = os.path.join(all_entities_folder, f"ALMM_ALL_All_Entities_{currency}.xlsx")
                self.save_to_excel(data, excel_file_path, global_file)

                # Sauvegarder par entité
                for entity in entity_list:
                    entity_data = data[data["Ref_Entite.entité"] == entity]
                    entity_folder = os.path.join(by_entity_folder, entity)
                    os.makedirs(entity_folder, exist_ok=True)

                    if not entity_data.empty:
                        entity_file = os.path.join(entity_folder, f"ALMM_ALL_{currency}_{entity}.xlsx")
                        self.save_to_excel(entity_data, excel_file_path, entity_file)
                        print(f"Fichier sauvegardé : {entity_file}")

    def save_to_excel(self, data: pd.DataFrame, template_path: str, output_path: str):
        """
        Sauvegarde des données dans un fichier Excel en utilisant un fichier template pour conserver la structure.

        :param data: DataFrame contenant les données à sauvegarder.
        :param template_path: Chemin du fichier Excel à utiliser comme template.
        :param output_path: Chemin du fichier Excel de sortie.
        """
        # Charger le classeur Excel existant
        workbook = load_workbook(template_path)
        first_sheet_name = workbook.sheetnames[0]  # Récupérer le nom de la première feuille
        first_sheet = workbook[first_sheet_name]  # Charger la première feuille uniquement

        # Effacer les anciennes données dans la première feuille
        for row in first_sheet.iter_rows():
            for cell in row:
                cell.value = None

        # Insérer les nouvelles données dans la première feuille
        for i, col_name in enumerate(data.columns, start=1):  # Parcourir les colonnes
            first_sheet.cell(row=1, column=i, value=col_name)  # Ajouter les noms de colonnes
            for j, value in enumerate(data[col_name], start=2):  # Parcourir les valeurs des colonnes
                first_sheet.cell(row=j, column=i, value=value)

        # Sauvegarder le fichier Excel avec les modifications
        workbook.save(output_path)
        print(f"Fichier sauvegardé : {output_path}")
