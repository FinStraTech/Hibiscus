import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import tempfile
import zipfile
import io

class LCR:
    def __init__(self, data_import: pd.DataFrame, ref_entite_path: str, ref_transfo_path: str, ref_lcr_path: str, ref_adf_lcr_path: str, input_excel_path: str, run_timestamp: str, export_type):


        self.data = data_import

        # Charger et prétraiter les fichiers de référence
        self.ref_entite = self.preprocess_ref_entite(ref_entite_path)
        self.ref_transfo = self.preprocess_ref_transfo(ref_transfo_path)
        self.ref_lcr = self.preprocess_ref_lcr(ref_lcr_path)
        self.ref_adf_lcr = self.preprocess_ref_adf_lcr(ref_adf_lcr_path)
        self.input_excel_path = input_excel_path
        self.run_timestamp = run_timestamp
        self.export_type = export_type

    
    
    def _save_import_files(self, filtered_data, import_folder, export_type):
        """
        Sauvegarde les fichiers d'import dans le dossier spécifié par devise (ALL, EUR, USD) 
        et vérifie que les fichiers générés ne sont pas corrompus.

        :param filtered_data: DataFrame filtré.
        :param import_folder: Dossier où sauvegarder les fichiers.
        :param export_type: Type d'export (ALL, BILAN, CONSO).
        :return: Dictionnaire contenant les chemins des fichiers générés.
        """
        saved_files = {}

        for currency in ["ALL", "EUR", "USD"]:
            if currency == "ALL":
                data_to_save = filtered_data
            else:
                data_to_save = filtered_data[filtered_data["D_CU"] == currency]

            # Vérifications avant sauvegarde
            if data_to_save.empty:
                print(f"Aucune donnée trouvée pour la devise {currency} dans {export_type}.")
                continue

            file_name = f"IMPORT_{export_type}_{currency}.xlsx"
            file_path = os.path.join(import_folder, file_name)

            try:
                # Étape 1 : Sauvegarder le fichier Excel
                data_to_save.to_excel(file_path, index=False, engine="xlsxwriter")
                print(f"Fichier généré : {file_path}")

                # Étape 2 : Valider que le fichier peut être relu correctement
                try:
                    test_read = pd.read_excel(file_path, engine="openpyxl")
                    if test_read.empty and not data_to_save.empty:
                        raise ValueError(f"Le fichier {file_path} est corrompu (lecture vide après écriture).")
                    if not data_to_save.equals(test_read):
                        raise ValueError(f"Le fichier {file_path} est corrompu (données lues non identiques à celles écrites).")
                except Exception as e:
                    raise ValueError(f"Validation échouée pour le fichier {file_path}: {e}")

                # Ajouter le fichier validé à la liste des fichiers sauvegardés
                saved_files[currency] = file_path

            except Exception as e:
                print(f"Erreur lors de la sauvegarde ou de la validation du fichier {file_path}: {e}")
                # Nettoyer le fichier corrompu s'il existe
                if os.path.exists(file_path):
                    os.remove(file_path)
                print(f"Fichier corrompu supprimé : {file_path}")

        return saved_files

    
    def preprocess_data(self, export_type="ALL", currency="ALL", entity="ALL"):
        """
        Nettoie et convertit les types des colonnes dans les données, génère les fichiers d'import
        pour BILAN, CONSO, ALL, et gère les étapes spécifiques pour GRAN.

        :param export_type: Type d'export choisi par l'utilisateur (ALL, BILAN, CONSO, GRAN).
        :param currency: Devise à filtrer (ALL, EUR, USD).
        :param entity: Entité à filtrer ou ALL.
        :return: Chemins des fichiers sauvegardés (dictionnaire) ou données filtrées (DataFrame) pour GRAN.
        """
        # Création du dossier d'import
        import_folder = f"./imports/import_{self.run_timestamp}"
        os.makedirs(import_folder, exist_ok=True)

        # Suppression des lignes totalement vides
        self.data = self.data.dropna(how="all")
        self.data = self.data[~self.data.apply(lambda row: all(row == ""), axis=1)]

        # Définition des types de colonnes
        column_types = {
            "D_CA": "string",
            "D_DP": "float64",
            "D_ZTFTR": "object",
            "D_PE": "float64",
            "D_RU": "string",
            "D_ORU": "string",
            "D_AC": "string",
            "D_FL": "string",
            "D_AU": "string",
            "D_T1": "object",
            "D_T2": "object",
            "D_CU": "string",
            "D_TO": "string",
            "D_GO": "string",
            "D_LE": "object",
            "D_NU": "object",
            "D_DEST": "object",
            "D_ZONE": "string",
            "D_MONNAIE": "string",
            "D_ENTITE": "object",
            "D_RESTIT": "object",
            "D_TYPCLI": "object",
            "D_SURFI": "object",
            "D_MU": "object",
            "D_PMU": "object",
            "D_ACTIVITE": "object",
            "D_ANALYSIS": "object",
            "D_PDT": "object",
            "P_AMOUNT": "Int64",
            "P_COMMENT": "object",
        }

        for col, dtype in column_types.items():
            if col in self.data.columns:
                try:
                    if dtype == "Int64":
                        self.data[col] = pd.to_numeric(self.data[col], errors='coerce').astype("Int64")
                    else:
                        self.data[col] = self.data[col].astype(dtype)
                except Exception as e:
                    print(f"Erreur lors de la conversion de la colonne {col} en {dtype}: {e}")

        # Étape 1 : Filtrage spécifique pour GRAN
        if export_type == "GRAN":
            if currency == "ALL":
                raise ValueError("Pour un export de type GRAN, une devise spécifique doit être fournie.")

            print(f"Filtrage des données pour la devise '{currency}'...")
            filtered_data_currency = self.data[self.data["D_CU"] == currency]

            if filtered_data_currency.empty:
                raise ValueError(f"Aucune donnée trouvée pour la devise '{currency}'.")

            return filtered_data_currency

        generated_files = {}
        #Étape 2 : Génération des fichiers pour BILAN, CONSO, et ALL
        generated_files = {}
        if export_type in ["ALL", "BILAN","CONSO","GRAN"]:
            filtered_bilan = self.data[self.data["D_T1"] == "INTER"]
            filtered_conso = self.data[self.data["D_T1"] != "INTER"]
            for save in self._save_import_files(filtered_bilan, "BILAN", import_folder, filtered_conso, "CONSO") :
                generated_files.update(save)
        print(f"Fichiers d'import sauvegardés dans : {import_folder}")
        return generated_files


    def _save_import_files(self, filtered_data_1, export_type_1, import_folder,filtered_data_2, export_type_2):
        """
        Sauvegarde les fichiers d'import dans le dossier spécifié par devise (ALL, EUR, USD).

        :param filtered_data: DataFrame filtré.
        :param import_folder: Dossier où sauvegarder les fichiers.
        :param export_type: Type d'export (ALL, BILAN, CONSO).
        :param all_data_accumulated: Liste pour accumuler les données.
        :return: Dictionnaire contenant les chemins des fichiers générés.
        """
        saved_files_1 = {}
        saved_files_2 = {}

        for currency in ["ALL", "EUR", "USD"]:
            if currency == "ALL":
                data_to_save_1 = filtered_data_1
                data_to_save_2 = filtered_data_2
            else:
                data_to_save_1 = filtered_data_1[filtered_data_1["D_CU"] == currency]
                data_to_save_2 = filtered_data_2[filtered_data_2["D_CU"] == currency]

            if data_to_save_1.empty or data_to_save_2.empty:
                print(f"Aucune donnée trouvée pour la devise {currency} dans {export_type}.")
                continue

            file_name_1 = f"IMPORT_{export_type_1}_{currency}.xlsx"
            file_name_2 = f"IMPORT_{export_type_2}_{currency}.xlsx"
            file_path_1 = os.path.join(import_folder, file_name_1)
            file_path_2 = os.path.join(import_folder, file_name_2)
            try:
                data_to_save_1.to_excel(file_path_1, index=False, engine="xlsxwriter")
                data_to_save_2.to_excel(file_path_2, index=False, engine="xlsxwriter")
                print(f"Fichier généré : {file_path_1} et {file_path_2}")
                saved_files_1[currency] = file_path_1
                saved_files_2[currency] = file_path_2
            except Exception as e:
                print(f"Erreur lors de la génération du fichier {file_path_1} ou {file_path_2}: {e}")

        return saved_files_1, saved_files_2




    def save_filtered_data(self, data: pd.DataFrame, file_name: str):
        """
        Sauvegarde les données filtrées dans un fichier Excel.

        :param data: DataFrame filtré à sauvegarder.
        :param file_name: Nom du fichier Excel de sortie.
        """
        file_path = f"./output/Exports/LCR/{file_name}" 
        os.makedirs(os.path.dirname(file_path), exist_ok=True) 
        data.to_excel(file_path, index=False, engine="openpyxl")
        print(f"Fichier sauvegardé : {file_path}")


    @staticmethod
    def preprocess_ref_entite(file_path: str) -> pd.DataFrame:
        """
        Prétraitement pour Ref_Entite.xlsx :
        - Supprime les lignes ayant une valeur nulle dans la colonne 'd_ru'.
        - Ajoute le préfixe 'Ref_Entite.' à tous les noms de colonnes.
        """
        df = pd.read_excel(file_path)
        df = df.dropna(subset=['d_ru'])  # Supprime les lignes où 'd_ru' est null

        # Renommer les colonnes en ajoutant le préfixe 'Ref_Entite.'
        df = df.rename(columns=lambda col: f"Ref_Entite.{col}")
        return df

    @staticmethod
    def preprocess_ref_transfo(file_path: str) -> pd.DataFrame:
        df = pd.read_excel(file_path)
        df['Transfo_aggregate_L1'] = df['Transfo_aggregate_L1'].astype(str)  # Convertit en texte
        df = df.drop_duplicates(subset=['Transfo_aggregate_L1'])  # Supprime les doublons

        # Renommer les colonnes en ajoutant le préfixe 'Ref_Transfo_L1.'
        df = df.rename(columns=lambda col: f"Ref_Transfo_L1.{col}")
        return df

    @staticmethod
    def preprocess_ref_lcr(file_path: str) -> pd.DataFrame:
        """
        Prétraitement pour Ref_LCR.xlsx :
        - Supprime les lignes où la valeur de 'Ligne_LCR' est nulle.
        - Ajoute le préfixe 'Ref_LCR.' à tous les noms de colonnes.
        """
        df = pd.read_excel(file_path)
        df = df.dropna(subset=['Ligne_LCR'])  # Supprime les lignes où 'Ligne_LCR' est null

        # Renommer les colonnes en ajoutant le préfixe 'Ref_LCR.'
        df = df.rename(columns=lambda col: f"Ref_LCR.{col}")
        return df

    @staticmethod
    def preprocess_ref_adf_lcr(file_path: str) -> pd.DataFrame:
        """
        Prétraitement pour Ref_ADF_LCR.xlsx :
        - Change les types des colonnes selon les spécifications.
        - Ajoute le préfixe 'Ref_ADF_LCR.' à tous les noms de colonnes.
        """
        df = pd.read_excel(file_path)
        column_types = {
            "D_ru": "string",
            "Entité": "string",
            "D_ac": "string",
            "Indicator_Ligne": "string",
            "Indicator_ADF": "Int64",
        }

        # Appliquer les conversions de types
        for col, dtype in column_types.items():
            if col in df.columns:
                try:
                    if dtype == "Int64":
                        df[col] = pd.to_numeric(df[col], errors='coerce').astype("Int64")
                    else:
                        df[col] = df[col].astype(dtype)
                except Exception as e:
                    print(f"Erreur lors de la conversion de la colonne {col} en {dtype}: {e}")

        # Renommer les colonnes en ajoutant le préfixe 'Ref_ADF_LCR.'
        df = df.rename(columns=lambda col: f"Ref_ADF_LCR.{col}")
        return df


    def filter_and_join_ref_entite(self,preprocessed_data):

        # 2.2. Filtrer les données
        filtered_data = preprocessed_data[
            (preprocessed_data["D_FL"] != "T99") & (preprocessed_data["D_ZONE"].notna())
        ]
        
        # 2.3. Joindre la table principale filtrée avec Ref_Entite
        joined_data = pd.merge(
            filtered_data,  # Table principale filtrée
            self.ref_entite,  # Table secondaire Ref_Entite
            left_on="D_RU",  # Colonne de jointure dans la table principale
            right_on="Ref_Entite.d_ru",  # Colonne de jointure dans la table secondaire
            how="left",  # Jointure externe gauche
        )

        # Retourner les données après jointure
        return joined_data
    
    def join_with_ref_transfo(self, filtered_data: pd.DataFrame):
        
        # Effectuer la jointure
        joined_data = pd.merge(
            filtered_data,  # Table principale (déjà filtrée et jointe avec Ref_Entite)
            self.ref_transfo,  # Référence Ref_Transfo_L1 (prétraitée dynamiquement)
            left_on="D_AC",  # Colonne de la table principale
            right_on="Ref_Transfo_L1.Transfo_aggregate_L1",  # Colonne de la référence
            how="left",  # Jointure externe gauche
        )

        # Filtrer les lignes où Transfo_aggregate_L1 n'est pas null
        filtered_joined_data = joined_data[joined_data["Ref_Transfo_L1.Transfo_aggregate_L1"].notna()]

        # Retourner les données après jointure et filtrage
        return filtered_joined_data

    def join_with_ref_lcr(self, filtered_data: pd.DataFrame):
        # Effectuer la jointure
        joined_data = pd.merge(
            filtered_data,  # Table principale
            self.ref_lcr,  # Référence Ref_LCR (prétraitée dynamiquement)
            left_on="D_AC",  # Colonne de la table principale
            right_on="Ref_LCR.Compte Transfo",  # Colonne de la référence
            how="left",  # Jointure externe gauche
        )

        # Retourner les données après jointure
        return joined_data
    
    def add_unadjusted_p_amount(self, data: pd.DataFrame) -> pd.DataFrame:

        # Vérifier que les colonnes nécessaires sont présentes
        required_columns = ["D_ZONE", "Ref_LCR.LCR_Flow_PCT", "Ref_LCR.LCR_Stock_PCT", "P_AMOUNT"]
        for col in required_columns:
            if col not in data.columns:
                raise ValueError(f"La colonne '{col}' est manquante dans le DataFrame.")

        # Ajouter la colonne 'Unadjusted_P_Amount'
        data["Unadjusted_P_Amount"] = data.apply(
            lambda row: row["Ref_LCR.LCR_Flow_PCT"] * row["P_AMOUNT"]
            if row["D_ZONE"] == "E01"
            else row["Ref_LCR.LCR_Stock_PCT"] * row["P_AMOUNT"],
            axis=1
        )

        return data

    
    def group_and_sum(self, data: pd.DataFrame):

        # Colonnes utilisées pour le regroupement
        group_columns = ["Ref_Entite.entité", "D_AC", "Ref_LCR.Ligne_LCR"]

        # Vérifier si les colonnes nécessaires sont présentes
        for col in group_columns + ["Unadjusted_P_Amount"]:
            if col not in data.columns:
                raise ValueError(f"La colonne '{col}' est manquante dans le DataFrame.")

        # Regrouper les données et calculer la somme
        grouped_data = (
            data.groupby(group_columns, as_index=False)
            .agg(Sum_Unadjusted_P_Amount=("Unadjusted_P_Amount", "sum"))
        )

        # Retourner le DataFrame regroupé
        return grouped_data

    def join_with_ref_adf_lcr(self, grouped_data: pd.DataFrame) -> pd.DataFrame:

        # Vérifier que les colonnes nécessaires sont présentes
        required_columns_main = ["D_AC", "Ref_LCR.Ligne_LCR"]
        required_columns_ref = ["Ref_ADF_LCR.D_ac", "Ref_ADF_LCR.Indicator_Ligne"]
        
        for col in required_columns_main:
            if col not in grouped_data.columns:
                raise ValueError(f"La colonne '{col}' est manquante dans le DataFrame principal.")

        for col in required_columns_ref:
            if col not in self.ref_adf_lcr.columns:
                raise ValueError(f"La colonne '{col}' est manquante dans Ref_ADF_LCR.")

        joined_data = pd.merge(
            grouped_data,  # Table principale après regroupement
            self.ref_adf_lcr,  # Référence Ref_ADF_LCR
            left_on=["D_AC", "Ref_LCR.Ligne_LCR"],  # Colonnes de jointure dans la table principale
            right_on=["Ref_ADF_LCR.D_ac", "Ref_ADF_LCR.Indicator_Ligne"],  # Colonnes de jointure dans la référence
            how="left", 
        )

        return joined_data

    def add_adjusted_amount(self, data: pd.DataFrame) -> pd.DataFrame:
        
        # Vérifier que les colonnes nécessaires sont présentes
        required_columns = ["Sum_Unadjusted_P_Amount", "Ref_ADF_LCR.Indicator_ADF"]
        for col in required_columns:
            if col not in data.columns:
                raise ValueError(f"La colonne '{col}' est manquante dans le DataFrame.")

        # Ajouter la colonne 'P_Adjusted_Amount'
        data["P_Adjusted_Amount"] = data["Sum_Unadjusted_P_Amount"] * data["Ref_ADF_LCR.Indicator_ADF"]

        return data
    

    def save_to_excel(self, data: pd.DataFrame, template_path: str, output_path: str, zip_buffer: zipfile.ZipFile):
        """
        Sauvegarde les données dans un fichier Excel en utilisant un fichier template, directement dans un ZIP.
        """
        # Charger le classeur Excel existant
        workbook = load_workbook(template_path)
        first_sheet_name = workbook.sheetnames[0]
        first_sheet = workbook[first_sheet_name]

        # Effacer les anciennes données
        for row in first_sheet.iter_rows():
            for cell in row:
                cell.value = None

        # Insérer les nouvelles données
        for i, col_name in enumerate(data.columns, start=1):
            first_sheet.cell(row=1, column=i, value=col_name)  # Ajouter les en-têtes
            for j, value in enumerate(data[col_name], start=2):
                first_sheet.cell(row=j, column=i, value=value)

        # Sauvegarder dans un fichier temporaire
        temp_file = io.BytesIO()
        workbook.save(temp_file)
        temp_file.seek(0)

        # Ajouter dans le ZIP
        zip_buffer.writestr(output_path, temp_file.getvalue())
        print(f"Fichier sauvegardé dans le ZIP : {output_path}")

    def save_excel_with_structure(
        self,
        processed_data: dict,
        template_path: str,
        entity_list: list,
        run_timestamp: str,
        export_type: str,
        zip_buffer: zipfile.ZipFile,
        entity: str = None,
        currency: str = "ALL"
    ):
        base_folder = f"RUN_{run_timestamp}_{export_type}"

        if not processed_data:
            st.warning("Aucune donnée à sauvegarder dans le ZIP.")
            return

        if export_type in ["BILAN", "CONSO"]:
            for currency, data in processed_data.items():
                if not isinstance(data, pd.DataFrame) or data.empty:
                    st.warning(f"Aucune donnée disponible pour la devise '{currency}'.")
                    continue

                global_file = f"{base_folder}/{export_type}_{currency}/Reports_all_entities/LCR_{export_type}_{currency}_All_Entities.xlsx"
                self.save_to_excel(data, template_path, global_file, zip_buffer)

                for entity in entity_list:
                    entity_data = data[data["Ref_Entite.entité"] == entity]
                    if not entity_data.empty:
                        entity_file = f"{base_folder}/{export_type}_{currency}/Reports_by_entity/{entity}/LCR_{export_type}_{currency}_{entity}.xlsx"
                        self.save_to_excel(entity_data, template_path, entity_file, zip_buffer)

        elif export_type == "ALL":
            for currency, data in processed_data.items():
                if not isinstance(data, pd.DataFrame) or data.empty:
                    st.warning(f"Aucune donnée disponible pour la devise '{currency}'.")
                    continue

                global_file = f"{base_folder}/Reports_all_entities/LCR_ALL_All_Entities_{currency}.xlsx"
                self.save_to_excel(data, template_path, global_file, zip_buffer)

                for entity in entity_list:
                    entity_data = data[data["Ref_Entite.entité"] == entity]
                    if not entity_data.empty:
                        entity_file = f"{base_folder}/Reports_by_entity/{entity}/LCR_ALL_{currency}_{entity}.xlsx"
                        self.save_to_excel(entity_data, template_path, entity_file, zip_buffer)
