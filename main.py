import pandas as pd
import os
import io
from concurrent.futures import ThreadPoolExecutor, as_completed
from LCR import LCR
from NSFR import NSFR
from AER import AER
from ALMM import ALMM
from QIS import QIS
from datetime import datetime
import streamlit as st
import shutil
import tempfile
import zipfile
from openpyxl import load_workbook
from io import BytesIO

Entity_List = ['BANCO SOCIETE GENERALE BRASIL SA','BPCE LEASE','FRAER LEASING SPA','FRANFINANCE','FRANFINANCE LOCATION','GEFA BANK GMBH','GERMAN NEWCO','GERMAN NEWCO','MILLA','PHILIPS MEDICAL CAPITAL FRANCE','SG EQUIPMENT FINANCE BENELUX BV','SG EQUIPMENT FINANCE CZECH REPUBLIC','SG EQUIPMENT FINANCE GMBH','SG EQUIPMENT FINANCE IBERIA','SG EQUIPMENT FINANCE ITALY SPA','SG EQUIPMENT FINANCE SCHWEIZ AG','SG EQUIPMENT FINANCE USA CORP','SG EQUIPMENT LEASING POLSKA SP ZO','SG EQUIPMENT LEASING POLSKA SP ZO','SG LEASING SPA','SGEF SA','SGEF SA ARRENDAMENTO MERCANTIL','SOCIETE GENERALE EQUIPMENT FINANCE Brazil','SOCIETE GENERALE EQUIPMENT FINANCE UK','SOCIETE GENERALE LEASING AND RENTING China']
expected_columns = [
    "D_CA", "D_DP", "D_ZTFTR", "D_PE", "D_RU", "D_ORU", "D_AC", "D_FL", "D_AU", 
    "D_T1", "D_T2", "D_CU", "D_TO", "D_GO", "D_LE", "D_NU", "D_DEST", "D_ZONE", 
    "D_MONNAIE", "D_ENTITE", "D_RESTIT", "D_TYPCLI", "D_SURFI", "D_MU", "D_PMU", 
    "D_ACTIVITE", "D_ANALYSIS", "D_PDT", "P_AMOUNT", "P_COMMENT"
]

def preprocess_all_data(data_path, ref_entite_path, ref_transfo_path, ref_lcr_path, ref_adf_lcr_path,
                        input_excel_path, run_timestamp, export_type, currency="ALL"):
    """
    Prétraitement des données pour tous les types d'export (ALL, BILAN, CONSO, GRAN).
    """
    try:
        data_import = pd.read_excel(data_path, engine="openpyxl")
    except Exception as e:
        raise ValueError(f"Erreur lors du chargement des données principales : {e}")

    # Vérifier les colonnes essentielles
    required_columns = ["D_CU", "D_T1", "D_ENTITE", "D_PE"]
    missing_columns = [col for col in required_columns if col not in data_import.columns]
    if missing_columns:
        raise ValueError(f"Les colonnes suivantes sont manquantes dans les données : {', '.join(missing_columns)}")

    # Initialiser le processeur LCR
    lcr_processor = LCR(
        data_import=data_import,
        ref_entite_path=ref_entite_path,
        ref_transfo_path=ref_transfo_path,
        ref_lcr_path=ref_lcr_path,
        ref_adf_lcr_path=ref_adf_lcr_path,
        input_excel_path=input_excel_path,
        run_timestamp=run_timestamp,
        export_type=export_type
    )

    # Prétraitement des données
    if export_type == "GRAN":
        # Vérification avant filtrage
        print(f"Valeurs uniques dans D_CU : {data_import['D_CU'].unique()}")

        # Filtrage des données par devise
        if currency == "ALL":
            filtered_data = data_import  # Prendre toutes les devises
        else:
            filtered_data = data_import[data_import["D_CU"] == currency]

        print("Filtrage réussi pour GRAN. Données disponibles :")
        print(filtered_data.head())  # Log des premières lignes pour vérification

        return {"filtered_data": filtered_data}
    else:
        # Prétraitement standard pour ALL, BILAN, et CONSO
        preprocessed_data = lcr_processor.preprocess_data(export_type=export_type, currency=currency)

        if isinstance(preprocessed_data, dict):
            return preprocessed_data
        else:
            raise ValueError("Le prétraitement des données a échoué pour les exports standard.")

def process_aer(preprocessed_data,
                data_path, ref_entite_path, ref_transfo_path, ref_aer_path, ref_adf_aer_path,
                input_excel_path, run_timestamp, export_type, zip_buffer,
                entity=None, currency=None, indicator="ALL"):
    """
    Processus pour traiter les données AER avec gestion spécifique des exports dans un ZIP,
    incluant la transition des données vers un fichier template.
    """
    base_folder = f"RUN_{run_timestamp}_{export_type}"  # Dossier racine dans le ZIP

    with zipfile.ZipFile(zip_buffer, 'a') as zipf:
        if export_type == "GRAN":

            if not entity or not currency:
                raise ValueError("Pour un export de type GRAN, une entité et une devise spécifiques doivent être fournies.")

            print(f"Traitement GRAN pour l'entité '{entity}' et la devise '{currency}'...")

            # Filtrer les données pour GRAN
            if isinstance(preprocessed_data, pd.DataFrame):
                if "D_CU" not in preprocessed_data.columns:
                    raise KeyError("La colonne 'D_CU' est absente dans les données prétraitées pour GRAN.")
                if currency == "ALL":
                    filtered_data = preprocessed_data
                else:
                    filtered_data = preprocessed_data[preprocessed_data["D_CU"] == currency]
            elif isinstance(preprocessed_data, dict):
                if "filtered_data" in preprocessed_data:
                    filtered_data = preprocessed_data["filtered_data"]
                    if "D_CU" not in filtered_data.columns:
                        raise KeyError("La colonne 'D_CU' est absente dans les données prétraitées pour GRAN.")
                    if currency == "ALL":
                        filtered_data = filtered_data
                    else:
                        filtered_data = filtered_data[filtered_data["D_CU"] == currency]
                else:
                    raise ValueError("La clé 'filtered_data' est absente dans preprocessed_data.")
            else:
                raise TypeError("preprocessed_data doit être un DataFrame ou un dictionnaire.")

            if filtered_data.empty:
                raise ValueError(f"Aucune donnée trouvée pour la devise '{currency}' dans l'export GRAN.")

            # Étape 2 : Filtrer par indicateur
            if indicator == "BILAN":
                filtered_data = filtered_data[filtered_data["D_T1"] == "INTER"]
            elif indicator == "CONSO":
                filtered_data = filtered_data[filtered_data["D_T1"] != "INTER"]
            elif indicator == "ALL":
                pass  # Ne rien filtrer
            else:
                raise ValueError("Indicateur non pris en charge. Choisissez parmi ALL, BILAN, ou CONSO.")

            if filtered_data.empty:
                raise ValueError(f"Aucune donnée trouvée pour l'indicateur '{indicator}'.")

            # Initialiser la classe AER
            aer_processor = AER(
                data_import=filtered_data,
                ref_entite_path=ref_entite_path,
                ref_transfo_path=ref_transfo_path,
                ref_aer_path=ref_aer_path,
                ref_adf_aer_path=ref_adf_aer_path,
                run_timestamp=run_timestamp,
                export_type=export_type,
            )

            # Appliquer les transformations
            result_after_entite = aer_processor.filter_and_join_ref_entite(filtered_data)
            result_after_transfo = aer_processor.join_with_ref_transfo(result_after_entite)
            result_with_aer = aer_processor.join_with_ref_aer(result_after_transfo)
            grouped_result = aer_processor.group_and_join_ref_adf_aer(result_with_aer)
            final_result = aer_processor.add_adjusted_amount(grouped_result)

            # Filtrer par entité
            final_result = final_result[final_result["Ref_Entite.entité"] == entity]

            # Transition vers le fichier template
            buffer = apply_to_template(final_result, input_excel_path)

            # Ajouter au ZIP
            folder_path = f"{base_folder}/{currency}/Reports_by_entity/{entity}"
            file_name = f"{folder_path}/AER_GRAN_{currency}_{entity}.xlsx"
            zipf.writestr(file_name, buffer.getvalue())

        else:  # Cas ALL, BILAN, CONSO
            for currency, file_path in preprocessed_data.items():
                if not os.path.exists(file_path):
                    print(f"Le fichier {file_path} n'existe pas. Aucun traitement pour cette devise.")
                    continue

                try:
                    data_import_filtered = pd.read_excel(file_path, engine="openpyxl")
                except Exception as e:
                    print(f"Erreur lors de la lecture du fichier {file_path}: {e}")
                    continue

                if data_import_filtered.empty:
                    continue

                print(f"Traitement de la devise : {currency}")

                # Initialiser la classe AER
                aer_processor = AER(
                    data_import=data_import_filtered,
                    ref_entite_path=ref_entite_path,
                    ref_transfo_path=ref_transfo_path,
                    ref_aer_path=ref_aer_path,
                    ref_adf_aer_path=ref_adf_aer_path,
                    run_timestamp=run_timestamp,
                    export_type=export_type,
                )

                # Appliquer les transformations
                result_after_entite = aer_processor.filter_and_join_ref_entite(data_import_filtered)
                result_after_transfo = aer_processor.join_with_ref_transfo(result_after_entite)
                result_with_aer = aer_processor.join_with_ref_aer(result_after_transfo)
                grouped_result = aer_processor.group_and_join_ref_adf_aer(result_with_aer)
                final_result = aer_processor.add_adjusted_amount(grouped_result)

                # Transition vers le fichier template
                buffer = apply_to_template(final_result, input_excel_path)

                # Ajouter au ZIP
                folder_path_global = f"{base_folder}/{currency}/Reports_all_entities"
                file_name_global = f"{folder_path_global}/AER_{export_type}_{currency}_All_Entities.xlsx"
                zipf.writestr(file_name_global, buffer.getvalue())
                
                # Ne générer que les rapports globaux si export_type == 'ALL'
                if export_type == 'ALL':
                    continue
                else:
                    # Sauvegarder les fichiers par entité
                    for entity in final_result["Ref_Entite.entité"].unique():
                        entity_data = final_result[final_result["Ref_Entite.entité"] == entity]
                        if entity_data.empty:
                            continue
                        buffer_entity = apply_to_template(entity_data, input_excel_path)
                        folder_path_entity = f"{base_folder}/{currency}/Reports_by_entity/{entity}"
                        file_name_entity = f"{folder_path_entity}/AER_{export_type}_{currency}_{entity}.xlsx"
                        zipf.writestr(file_name_entity, buffer_entity.getvalue())

    print("Tous les fichiers AER ont été ajoutés au ZIP.")


def process_qis(
    preprocessed_data,
    data_path,
    ref_entite_path,
    ref_transfo_path,
    ref_qis_path,
    ref_adf_qis_path,
    ref_dzone_qis_path,
    input_excel_path,
    run_timestamp,
    export_type,
    zip_buffer,
    entity=None,
    currency=None,
    indicator="ALL"
):
    """
    Processus pour traiter les données QIS avec gestion spécifique des exports dans un ZIP,
    incluant la transition des données vers un fichier template.
    """
    base_folder = f"RUN_{run_timestamp}_{export_type}"  # Dossier racine dans le ZIP

    with zipfile.ZipFile(zip_buffer, 'a') as zipf:
        if export_type == "GRAN":
            if not entity or not currency:
                raise ValueError("Pour un export de type GRAN, une entité et une devise spécifiques doivent être fournies.")

            print(f"Traitement GRAN pour l'entité '{entity}' et la devise '{currency}'...")

            # Filtrer les données pour GRAN
            if isinstance(preprocessed_data, pd.DataFrame):
                if "D_CU" not in preprocessed_data.columns:
                    raise KeyError("La colonne 'D_CU' est absente dans les données prétraitées pour GRAN.")
                if currency == "ALL":
                    filtered_data = preprocessed_data
                else:
                    filtered_data = preprocessed_data[preprocessed_data["D_CU"] == currency]
            else:
                raise TypeError("preprocessed_data doit être un DataFrame pour un export de type GRAN.")

            if filtered_data.empty:
                raise ValueError(f"Aucune donnée trouvée pour la devise '{currency}' dans l'export GRAN.")

            # Initialiser la classe QIS
            qis_processor = QIS(
                data_import=filtered_data,
                ref_entite_path=ref_entite_path,
                ref_transfo_path=ref_transfo_path,
                ref_qis_path=ref_qis_path,
                ref_adf_qis_path=ref_adf_qis_path,
                ref_dzone_qis_path=ref_dzone_qis_path,
                run_timestamp=run_timestamp,
                export_type=export_type,
            )

            # Appliquer les transformations
            result_after_entite = qis_processor.filter_and_join_ref_entite(filtered_data)
            result_after_transfo = qis_processor.join_with_ref_transfo(result_after_entite)
            result_with_qis = qis_processor.join_with_ref_qis(result_after_transfo)
            grouped_result = qis_processor.group_and_join_ref_adf_qis(result_with_qis)
            final_result = qis_processor.add_adjusted_amount(grouped_result)

            # Filtrer par entité
            final_result = final_result[final_result["Ref_Entite.entité"] == entity]

            # Transition vers le fichier template
            buffer = apply_to_template(final_result, input_excel_path)

            # Ajouter au ZIP
            folder_path = f"{base_folder}/{currency}/Reports_by_entity/{entity}"
            file_name = f"{folder_path}/QIS_GRAN_{currency}_{entity}.xlsx"
            zipf.writestr(file_name, buffer.getvalue())

        else:  # Cas ALL, BILAN, CONSO
            for currency, file_path in preprocessed_data.items():
                if not os.path.exists(file_path):
                    print(f"Le fichier {file_path} n'existe pas. Aucun traitement pour cette devise.")
                    continue

                try:
                    data_import_filtered = pd.read_excel(file_path, engine="openpyxl")
                except Exception as e:
                    print(f"Erreur lors de la lecture du fichier {file_path}: {e}")
                    continue

                if data_import_filtered.empty:
                    continue

                print(f"Traitement de la devise : {currency}")

                # Initialiser la classe QIS
                qis_processor = QIS(
                    data_import=data_import_filtered,
                    ref_entite_path=ref_entite_path,
                    ref_transfo_path=ref_transfo_path,
                    ref_qis_path=ref_qis_path,
                    ref_adf_qis_path=ref_adf_qis_path,
                    ref_dzone_qis_path=ref_dzone_qis_path,
                    run_timestamp=run_timestamp,
                    export_type=export_type,
                )

                result_after_entite = qis_processor.filter_and_join_ref_entite(data_import_filtered)
                result_after_transfo = qis_processor.join_with_ref_transfo(result_after_entite)
                result_with_dzone_qis = qis_processor.join_with_ref_dzone_qis(result_after_transfo)
                result_with_qis = qis_processor.join_with_ref_qis(result_with_dzone_qis)
                grouped_result = qis_processor.group_and_sum_unadjusted_p_amount(result_with_qis)
                pivoted_and_reordered_result = qis_processor.pivot_and_reorder(grouped_result)
                final_result_with_adf_qis = qis_processor.join_with_ref_adf_qis(pivoted_and_reordered_result)
                final_result = qis_processor.add_adjusted_amounts(final_result_with_adf_qis)

                # Transition vers le fichier template
                buffer = apply_to_template(final_result, input_excel_path)

                # Ajouter au ZIP
                folder_path_global = f"{base_folder}/{currency}/Reports_all_entities"
                file_name_global = f"{folder_path_global}/QIS_{export_type}_{currency}_All_Entities.xlsx"
                zipf.writestr(file_name_global, buffer.getvalue())

                # Ne générer que les rapports globaux si export_type == 'ALL'
                if export_type == 'ALL':
                    continue
                else:
                    # Sauvegarder les fichiers par entité
                    for entity in final_result["Ref_Entite.entité"].unique():
                        entity_data = final_result[final_result["Ref_Entite.entité"] == entity]
                        if entity_data.empty:
                            continue
                        buffer_entity = apply_to_template(entity_data, input_excel_path)
                        folder_path_entity = f"{base_folder}/{currency}/Reports_by_entity/{entity}"
                        file_name_entity = f"{folder_path_entity}/QIS_{export_type}_{currency}_{entity}.xlsx"
                        zipf.writestr(file_name_entity, buffer_entity.getvalue())

    print("Tous les fichiers QIS ont été ajoutés au ZIP.")

def process_almm(preprocessed_data,
    data_path, ref_entite_path, ref_transfo_path, ref_almm_path, ref_adf_almm_path,
    ref_dzone_almm_path, input_excel_path, run_timestamp, export_type, zip_buffer, entity=None, currency=None, indicator="ALL"
):
    """
    Processus pour traiter les données ALMM avec gestion spécifique des exports dans un ZIP.
    """
    base_folder = f"RUN_{run_timestamp}_{export_type}"  # Dossier racine dans le ZIP

    with zipfile.ZipFile(zip_buffer, 'a') as zipf:
        if export_type == "GRAN":
            if not entity or not currency:
                raise ValueError("Pour un export de type GRAN, une entité et une devise spécifiques doivent être fournies.")

            print(f"Traitement GRAN pour l'entité '{entity}' et la devise '{currency}'...")

            # Filtrer les données pour GRAN
            if isinstance(preprocessed_data, pd.DataFrame):
                # Si c'est un DataFrame, afficher ses colonnes
                if "D_CU" not in preprocessed_data.columns:
                    raise KeyError("La colonne 'D_CU' est absente dans les données prétraitées pour GRAN.")
                if currency == "ALL":
                    filtered_data = preprocessed_data
                else:
                    filtered_data = preprocessed_data[preprocessed_data["D_CU"] == currency]
            elif isinstance(preprocessed_data, dict):
                # Si c'est un dictionnaire, accéder à la clé "filtered_data"
                if "filtered_data" in preprocessed_data:
                    filtered_data = preprocessed_data["filtered_data"]
                    if "D_CU" not in filtered_data.columns:
                        raise KeyError("La colonne 'D_CU' est absente dans les données prétraitées pour GRAN.")
                    if currency == "ALL":
                        filtered_data = filtered_data
                    else:
                        filtered_data = filtered_data[filtered_data["D_CU"] == currency]
                else:
                    raise ValueError("La clé 'filtered_data' est absente dans preprocessed_lcr_data.")
            else:
                raise TypeError("preprocessed_lcr_data doit être un DataFrame ou un dictionnaire.")

            # Vérifier si 'filtered_data' est valide
            if filtered_data.empty:
                st.error(f"Aucune donnée trouvée pour la devise '{currency}' dans l'export GRAN.")


            # Étape 2 : Filtrer par indicateur
            if indicator == "BILAN":
                filtered_data = filtered_data[filtered_data["D_T1"] == "INTER"]
            elif indicator == "CONSO":
                filtered_data = filtered_data[filtered_data["D_T1"] != "INTER"]
            elif indicator == "ALL":
                filtered_data = filtered_data
            else:
                raise ValueError("Indicateur non pris en charge. Choisissez parmi ALL, BILAN, ou CONSO.")

            if filtered_data.empty:
                raise ValueError(f"Aucune donnée trouvée pour l'indicateur '{indicator}'.")

            # Initialiser la classe ALMM
            almm_processor = ALMM(
                data_import=filtered_data,
                ref_entite_path=ref_entite_path,
                ref_transfo_path=ref_transfo_path,
                ref_almm_path=ref_almm_path,
                ref_adf_almm_path=ref_adf_almm_path,
                ref_dzone_almm_path=ref_dzone_almm_path,
                run_timestamp=run_timestamp,
                export_type=export_type,
            )

            # Appliquer les transformations
            result_after_entite = almm_processor.filter_and_join_ref_entite(filtered_data)
            result_after_transfo = almm_processor.join_with_ref_transfo(result_after_entite)
            result_with_dzone_almm = almm_processor.join_with_ref_dzone_almm(result_after_transfo)
            result_with_almm = almm_processor.join_with_ref_almm(result_with_dzone_almm)
            grouped_result = almm_processor.group_and_sum_unadjusted_p_amount(result_with_almm)
            pivoted_and_reordered_result = almm_processor.pivot_and_reorder(grouped_result)
            final_result_with_adf_almm = almm_processor.join_with_ref_adf_almm(pivoted_and_reordered_result)
            final_result = almm_processor.add_adjusted_amounts(final_result_with_adf_almm)

            # Filtrer par entité
            final_result = final_result[final_result["Ref_Entite.entité"] == entity]

            # Sauvegarder dans le ZIP
            folder_path = f"{base_folder}/{currency}/Reports_by_entity/{entity}"
            file_name = f"{folder_path}/ALMM_GRAN_{currency}_{entity}.xlsx"
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_file_path = os.path.join(temp_dir, "temp_output.xlsx")
                try:
                    final_result.to_excel(temp_file_path, index=False, engine="xlsxwriter")
                    zipf.write(temp_file_path, arcname=file_name)
                except PermissionError as e:
                    print(f"Erreur de permission lors de la création du fichier : {e}")
                except Exception as e:
                    print(f"Une erreur inattendue s'est produite : {e}")

        else:  # Cas ALL, BILAN, CONSO
            for currency, file_path in preprocessed_data.items():
                if not os.path.exists(file_path):
                    print(f"Le fichier {file_path} n'existe pas. Aucun traitement pour cette devise.")
                    continue

                try:
                    data_import_filtered = pd.read_excel(file_path, engine="openpyxl")
                except Exception as e:
                    print(f"Erreur lors de la lecture du fichier {file_path}: {e}")
                    continue

                if data_import_filtered.empty:
                    continue

                print(f"Traitement de la devise : {currency}")

                # Initialiser la classe ALMM
                almm_processor = ALMM(
                    data_import=data_import_filtered,
                    ref_entite_path=ref_entite_path,
                    ref_transfo_path=ref_transfo_path,
                    ref_almm_path=ref_almm_path,
                    ref_adf_almm_path=ref_adf_almm_path,
                    ref_dzone_almm_path=ref_dzone_almm_path,
                    run_timestamp=run_timestamp,
                    export_type=export_type,
                )

                # Appliquer les transformations
                result_after_entite = almm_processor.filter_and_join_ref_entite(data_import_filtered)
                result_after_transfo = almm_processor.join_with_ref_transfo(result_after_entite)
                result_with_dzone_almm = almm_processor.join_with_ref_dzone_almm(result_after_transfo)
                result_with_almm = almm_processor.join_with_ref_almm(result_with_dzone_almm)
                grouped_result = almm_processor.group_and_sum_unadjusted_p_amount(result_with_almm)
                pivoted_and_reordered_result = almm_processor.pivot_and_reorder(grouped_result)
                final_result_with_adf_almm = almm_processor.join_with_ref_adf_almm(pivoted_and_reordered_result)
                final_result = almm_processor.add_adjusted_amounts(final_result_with_adf_almm)

                # Sauvegarder le fichier global
                folder_path_global = f"{base_folder}/{currency}/Reports_all_entities"
                file_name_global = f"{folder_path_global}/ALMM_{export_type}_{currency}_All_Entities.xlsx"
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                    final_result.to_excel(temp_file.name, index=False, engine="xlsxwriter")
                    zipf.write(temp_file.name, arcname=file_name_global)
                
                # Ne générer que les rapports globaux si export_type == 'ALL'
                if export_type == 'ALL':
                    continue
                else:
                    # Sauvegarder les fichiers par entité
                    for entity in final_result["Ref_Entite.entité"].unique():
                        entity_data = final_result[final_result["Ref_Entite.entité"] == entity]
                        if entity_data.empty:
                            continue
                        folder_path_entity = f"{base_folder}/{currency}/Reports_by_entity/{entity}"
                        file_name_entity = f"{folder_path_entity}/ALMM_{export_type}_{currency}_{entity}.xlsx"
                        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                            entity_data.to_excel(temp_file.name, index=False, engine="xlsxwriter")
                            zipf.write(temp_file.name, arcname=file_name_entity)

    print("Tous les fichiers ALMM ont été ajoutés au ZIP.")


def process_nsfr(preprocessed_data,
                 data_path, ref_entite_path, ref_transfo_path, ref_nsfr_path, ref_adf_nsfr_path, ref_dzone_nsfr_path,
                 input_excel_path, run_timestamp, export_type, zip_buffer, entity=None, currency=None, indicator="ALL"):
    """
    Processus de traitement des données NSFR avec intégration des résultats dans un fichier template
    et gestion des exports structurés dans un ZIP.
    """
    if zip_buffer is None:
        raise ValueError("Le buffer ZIP n'est pas initialisé.")

    base_folder = f"RUN_{run_timestamp}_{export_type}"  # Dossier racine dans le ZIP

    with zipfile.ZipFile(zip_buffer, 'a') as zipf:
        if export_type == "GRAN":
            if not entity or not currency:
                raise ValueError("Pour un export de type GRAN, une entité et une devise spécifiques doivent être fournies.")

            print(f"Traitement GRAN pour l'entité '{entity}' et la devise '{currency}'...")

            # Filtrer les données pour GRAN
            if isinstance(preprocessed_data, pd.DataFrame):
                if "D_CU" not in preprocessed_data.columns:
                    raise KeyError("La colonne 'D_CU' est absente dans les données prétraitées pour GRAN.")
                if currency == "ALL":
                    filtered_data = preprocessed_data
                else:
                    filtered_data = preprocessed_data[preprocessed_data["D_CU"] == currency]
            elif isinstance(preprocessed_data, dict):
                if "filtered_data" in preprocessed_data:
                    filtered_data = preprocessed_data["filtered_data"]
                    if "D_CU" not in filtered_data.columns:
                        raise KeyError("La colonne 'D_CU' est absente dans les données prétraitées pour GRAN.")
                    if currency == "ALL":
                        filtered_data = filtered_data
                    else:
                        filtered_data = filtered_data[filtered_data["D_CU"] == currency]
                else:
                    raise ValueError("La clé 'filtered_data' est absente dans preprocessed_data.")
            else:
                raise TypeError("preprocessed_data doit être un DataFrame ou un dictionnaire.")

            if filtered_data.empty:
                raise ValueError(f"Aucune donnée trouvée pour la devise '{currency}' dans l'export GRAN.")

            # Étape 2 : Filtrer par indicateur
            if indicator == "BILAN":
                filtered_data = filtered_data[filtered_data["D_T1"] == "INTER"]
            elif indicator == "CONSO":
                filtered_data = filtered_data[filtered_data["D_T1"] != "INTER"]
            elif indicator != "ALL":
                raise ValueError("Indicateur non pris en charge. Choisissez parmi ALL, BILAN, ou CONSO.")

            if filtered_data.empty:
                raise ValueError(f"Aucune donnée trouvée pour l'indicateur '{indicator}'.")

            # Initialiser le processeur NSFR
            nsfr_processor = NSFR(
                data_import=filtered_data,
                ref_entite_path=ref_entite_path,
                ref_transfo_path=ref_transfo_path,
                ref_nsfr_path=ref_nsfr_path,
                ref_adf_nsfr_path=ref_adf_nsfr_path,
                ref_dzone_nsfr_path=ref_dzone_nsfr_path,
                run_timestamp=run_timestamp,
                export_type=export_type,
            )

            # Étapes de transformation
            result_after_entite = nsfr_processor.filter_and_join_ref_entite(filtered_data)
            result_after_transfo = nsfr_processor.join_with_ref_transfo(result_after_entite)
            result_with_dzone_nsfr = nsfr_processor.join_with_ref_dzone_nsfr(result_after_transfo)
            result_with_nsfr = nsfr_processor.join_with_ref_nsfr(result_with_dzone_nsfr)
            grouped_result = nsfr_processor.group_and_sum_unadjusted_p_amount(result_with_nsfr)
            pivoted_and_reordered_result = nsfr_processor.pivot_and_reorder(grouped_result)
            final_result_with_adf_nsfr = nsfr_processor.join_with_ref_adf_nsfr(pivoted_and_reordered_result)
            final_result = nsfr_processor.add_adjusted_amounts(final_result_with_adf_nsfr)

            # Filtrer par entité
            final_result = final_result[final_result["Ref_Entite.entité"] == entity]

            # Transition vers le fichier template
            buffer = apply_to_template(final_result, input_excel_path)

            # Ajouter au ZIP
            folder_path = f"{base_folder}/{currency}/Reports_by_entity/{entity}"
            file_name = f"{folder_path}/NSFR_GRAN_{currency}_{entity}.xlsx"
            zipf.writestr(file_name, buffer.getvalue())

        else:  # Cas ALL, BILAN, CONSO
            for currency, file_path in preprocessed_data.items():
                if not os.path.exists(file_path):
                    print(f"Le fichier {file_path} n'existe pas. Aucun traitement pour cette devise.")
                    continue

                try:
                    data_import_filtered = pd.read_excel(file_path, engine="openpyxl")
                except Exception as e:
                    print(f"Erreur lors de la lecture du fichier {file_path}: {e}")
                    continue

                if data_import_filtered.empty:
                    continue

                print(f"Traitement de la devise : {currency}")

                nsfr_processor = NSFR(
                    data_import=data_import_filtered,
                    ref_entite_path=ref_entite_path,
                    ref_transfo_path=ref_transfo_path,
                    ref_nsfr_path=ref_nsfr_path,
                    ref_adf_nsfr_path=ref_adf_nsfr_path,
                    ref_dzone_nsfr_path=ref_dzone_nsfr_path,
                    run_timestamp=run_timestamp,
                    export_type=export_type,
                )

                # Étapes de transformation
                result_after_entite = nsfr_processor.filter_and_join_ref_entite(data_import_filtered)
                result_after_transfo = nsfr_processor.join_with_ref_transfo(result_after_entite)
                result_with_dzone_nsfr = nsfr_processor.join_with_ref_dzone_nsfr(result_after_transfo)
                result_with_nsfr = nsfr_processor.join_with_ref_nsfr(result_with_dzone_nsfr)
                grouped_result = nsfr_processor.group_and_sum_unadjusted_p_amount(result_with_nsfr)
                pivoted_and_reordered_result = nsfr_processor.pivot_and_reorder(grouped_result)
                final_result_with_adf_nsfr = nsfr_processor.join_with_ref_adf_nsfr(pivoted_and_reordered_result)
                final_result = nsfr_processor.add_adjusted_amounts(final_result_with_adf_nsfr)

                # Transition vers le fichier template global
                buffer = apply_to_template(final_result, input_excel_path)

                # Ajouter au ZIP
                folder_path = f"{base_folder}/{currency}/Reports_all_entities"
                file_name = f"{folder_path}/NSFR_{export_type}_{currency}_All_Entities.xlsx"
                zipf.writestr(file_name, buffer.getvalue())
                
                # Ne générer que les rapports globaux si export_type == 'ALL'
                if export_type == 'ALL':
                    continue
                
                
                else:
                    # Sauvegarder les fichiers par entité
                    for entity in final_result["Ref_Entite.entité"].unique():
                        entity_data = final_result[final_result["Ref_Entite.entité"] == entity]
                        if entity_data.empty:
                            continue
                        buffer_entity = apply_to_template(entity_data, input_excel_path)
                        folder_path_entity = f"{base_folder}/{currency}/Reports_by_entity/{entity}"
                        file_name_entity = f"{folder_path_entity}/NSFR_{export_type}_{currency}_{entity}.xlsx"
                        zipf.writestr(file_name_entity, buffer_entity.getvalue())

    print("Tous les fichiers NSFR ont été ajoutés au ZIP.")



def process_lcr(preprocessed_lcr_data,
                data_path, ref_entite_path, ref_transfo_path, ref_lcr_path, ref_adf_lcr_path,
                input_excel_path, run_timestamp, export_type, zip_buffer, entity=None, currency=None, indicator="ALL"):
    """
    Processus de traitement des données LCR avec transition directe des données dans un fichier template
    et stockage des fichiers générés dans un ZIP en mémoire.
    """
    base_folder = f"RUN_{run_timestamp}_{export_type}"  # Dossier racine dans le ZIP

    with zipfile.ZipFile(zip_buffer, 'a') as zipf:
        if export_type == "GRAN":
            if not entity or not currency:
                raise ValueError("Pour un export de type GRAN, une entité et une devise spécifiques doivent être fournies.")

            print(f"Traitement GRAN pour l'entité '{entity}' et la devise '{currency}'...")

            # Filtrer les données pour GRAN
            if isinstance(preprocessed_lcr_data, pd.DataFrame):
                if "D_CU" not in preprocessed_lcr_data.columns:
                    raise KeyError("La colonne 'D_CU' est absente dans les données prétraitées pour GRAN.")
                if currency == "ALL":
                    filtered_data = preprocessed_lcr_data
                else:
                    filtered_data = preprocessed_lcr_data[preprocessed_lcr_data["D_CU"] == currency]
            elif isinstance(preprocessed_lcr_data, dict):
                if "filtered_data" in preprocessed_lcr_data:
                    filtered_data = preprocessed_lcr_data["filtered_data"]
                    if "D_CU" not in filtered_data.columns:
                        raise KeyError("La colonne 'D_CU' est absente dans les données prétraitées pour GRAN.")
                    if currency == "ALL":
                        filtered_data = filtered_data
                    else:
                        filtered_data = filtered_data[filtered_data["D_CU"] == currency]
                else:
                    raise ValueError("La clé 'filtered_data' est absente dans preprocessed_lcr_data.")
            else:
                raise TypeError("preprocessed_lcr_data doit être un DataFrame ou un dictionnaire.")

            # Vérification des données filtrées
            if filtered_data.empty:
                print(f"Attention : aucune donnée trouvée pour la devise '{currency}' avec export GRAN.")
                return

            # Initialiser le processeur LCR
            lcr_processor = LCR(
                data_import=filtered_data,
                ref_entite_path=ref_entite_path,
                ref_transfo_path=ref_transfo_path,
                ref_lcr_path=ref_lcr_path,
                ref_adf_lcr_path=ref_adf_lcr_path,
                input_excel_path=input_excel_path,
                run_timestamp=run_timestamp,
                export_type=export_type,
            )

            # Étapes de transformation
            result_after_entite = lcr_processor.filter_and_join_ref_entite(filtered_data)
            result_after_transfo = lcr_processor.join_with_ref_transfo(result_after_entite)
            result_after_lcr = lcr_processor.join_with_ref_lcr(result_after_transfo)
            result_with_amount = lcr_processor.add_unadjusted_p_amount(result_after_lcr)
            grouped_result = lcr_processor.group_and_sum(result_with_amount)
            result_with_adf = lcr_processor.join_with_ref_adf_lcr(grouped_result)
            final_result = lcr_processor.add_adjusted_amount(result_with_adf)
            final_result = final_result[final_result["Ref_Entite.entité"] == entity]

            # Vérification des données finales
            if final_result.empty:
                print(f"Aucune donnée à exporter pour l'entité '{entity}' et la devise '{currency}'.")
                return

            # Transition vers le fichier template
            buffer = apply_to_template(final_result, input_excel_path)

            if buffer.getvalue() == b"":
                print("Le buffer est vide ! Vérifiez la fonction apply_to_template.")
                return

            # Ajouter au ZIP
            folder_path = f"{base_folder}/{currency}/Reports_by_entity/{entity}"
            file_name = f"{folder_path}/LCR_GRAN_{currency}_{entity}.xlsx"
            print(f"Écriture dans le ZIP : {file_name}")
            zipf.writestr(file_name, buffer.getvalue())

        else:  # Pour ALL, BILAN, CONSO
            for currency, filtered_data in preprocessed_lcr_data.items():
                if isinstance(filtered_data, str):
                    try:
                        filtered_data = pd.read_excel(filtered_data, engine="openpyxl")
                    except Exception as e:
                        print(f"Erreur lors de la lecture du fichier {filtered_data}: {e}")
                        continue

                if filtered_data.empty:
                    continue

                print(f"Traitement de la devise : {currency}")

                # Initialiser le processeur LCR
                lcr_processor = LCR(
                    data_import=filtered_data,
                    ref_entite_path=ref_entite_path,
                    ref_transfo_path=ref_transfo_path,
                    ref_lcr_path=ref_lcr_path,
                    ref_adf_lcr_path=ref_adf_lcr_path,
                    input_excel_path=input_excel_path,
                    run_timestamp=run_timestamp,
                    export_type=export_type,
                )

                # Transformation des données
                result_after_entite = lcr_processor.filter_and_join_ref_entite(filtered_data)
                result_after_transfo = lcr_processor.join_with_ref_transfo(result_after_entite)
                result_after_lcr = lcr_processor.join_with_ref_lcr(result_after_transfo)
                result_with_amount = lcr_processor.add_unadjusted_p_amount(result_after_lcr)
                grouped_result = lcr_processor.group_and_sum(result_with_amount)
                result_with_adf = lcr_processor.join_with_ref_adf_lcr(grouped_result)
                final_result = lcr_processor.add_adjusted_amount(result_with_adf)


                # Transition vers le fichier template global
                buffer = apply_to_template(final_result, input_excel_path)

                # Ajouter au ZIP
                folder_path_global = f"{base_folder}/{currency}/Reports_all_entities"
                file_name_global = f"{folder_path_global}/LCR_{export_type}_{currency}_All_Entities.xlsx"
                with zipfile.ZipFile(zip_buffer, "a") as zipf:
                    zipf.writestr(file_name_global, buffer.getvalue())
                    
                # Ne générer que les rapports globaux si export_type == 'ALL'
                if export_type == 'ALL':
                    continue
                else:
                    # Sauvegarder les fichiers par entité
                    for entity in final_result["Ref_Entite.entité"].unique():
                        entity_data = final_result[final_result["Ref_Entite.entité"] == entity]
                        if entity_data.empty:
                            continue
                        buffer_entity = apply_to_template(entity_data, input_excel_path)
                        folder_path_entity = f"{base_folder}/{currency}/Reports_by_entity/{entity}"
                        file_name_entity = f"{folder_path_entity}/LCR_{export_type}_{currency}_{entity}.xlsx"
                        with zipfile.ZipFile(zip_buffer, "a") as zipf:
                            zipf.writestr(file_name_entity, buffer_entity.getvalue())


def apply_to_template(dataframe, template_path):
    """
    Applique les données d'un DataFrame dans un fichier de template.
    Les colonnes du DataFrame doivent correspondre exactement à celles du template.
    
    :param dataframe: DataFrame contenant les données à insérer.
    :param template_path: Chemin du fichier Excel template.
    :return: Un buffer contenant le fichier Excel modifié.
    """
    buffer = BytesIO()
    
    # Charger le template
    workbook = load_workbook(template_path)
    sheet = workbook.active  # Utiliser la première feuille

    # Effacer les données existantes dans la feuille (à partir de la 2e ligne)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.value = None

    # Insérer les données du DataFrame dans le template
    for i, row in enumerate(dataframe.values, start=2):  # Commence à la ligne 2
        for j, value in enumerate(row, start=1):  # Commence à la colonne 1
            sheet.cell(row=i, column=j, value=value)

    # Sauvegarder dans un buffer en mémoire
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

def add_file_to_zip(zip_buffer, file_path, arcname):
    """
    Ajoute un fichier au buffer ZIP avec une gestion des erreurs.

    :param zip_buffer: Buffer ZIP en mémoire.
    :param file_path: Chemin absolu du fichier à ajouter.
    :param arcname: Nom du fichier à l'intérieur du ZIP.
    """
    try:
        with zipfile.ZipFile(zip_buffer, mode="a") as zipf:
            zipf.write(file_path, arcname=arcname)
    except Exception as e:
        raise RuntimeError(f"Erreur lors de l'ajout du fichier {file_path} au ZIP : {e}")
                            
def validate_zip_content(zip_buffer, expected_files):
    """
    Valide que tous les fichiers attendus sont dans le buffer ZIP.

    :param zip_buffer: Buffer ZIP en mémoire.
    :param expected_files: Liste des chemins attendus à l'intérieur du ZIP.
    """
    with zipfile.ZipFile(zip_buffer, 'r') as zipf:
        zip_contents = zipf.namelist()
        missing_files = [file for file in expected_files if file not in zip_contents]
        if missing_files:
            raise ValueError(f"Les fichiers suivants manquent dans le ZIP : {missing_files}")


def execute_processes_in_parallel(processes):
    """
    Exécute plusieurs processus en parallèle.

    :param processes: Liste de tuples contenant une fonction à exécuter et ses arguments.
    Format : [(fonction, (arg1, arg2, ...)), ...]
    :return: Résultats et erreurs des processus.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    
    results = {}
    errors = {}

    with ThreadPoolExecutor() as executor:
        # Soumettre toutes les tâches
        future_to_process = {
            executor.submit(func, *args): func.__name__ for func, args in processes
        }

        for future in as_completed(future_to_process):
            func_name = future_to_process[future]
            try:
                result = future.result()  # Récupérer le résultat de la fonction
                results[func_name] = result
            except Exception as e:
                errors[func_name] = str(e)  # Capturer l'exception

    return results, errors


def save_hierarchy_to_excel_from_directory(base_dir, output_file):
    """
    Sauvegarde la hiérarchie des fichiers et dossiers dans un fichier Excel.
    """
    def extract_hierarchy_from_paths(base_dir):
        hierarchy = {}
        seen_files = set()
        for root, dirs, files in os.walk(base_dir):
            relative_path = os.path.relpath(root, base_dir)
            levels = relative_path.split(os.sep) if relative_path != "." else []
            current_level = hierarchy
            for level in levels:
                current_level = current_level.setdefault(level, {})
            for file in files:
                file_path = os.path.join(relative_path, file)
                if file_path not in seen_files:
                    current_level[file] = None
                    seen_files.add(file_path)
        return hierarchy
    
    hierarchy = extract_hierarchy_from_paths(base_dir)
    rows = []

    def traverse_hierarchy(parent, structure, level=0):
        if isinstance(structure, dict):
            for key, value in structure.items():
                rows.append((level, key))
                traverse_hierarchy(key, value, level + 1)
        elif structure is None:
            rows.append((level, parent))

    traverse_hierarchy(None, hierarchy)

    if not rows:
        print(f"Aucune hiérarchie trouvée dans le dossier : {base_dir}")
        return

    max_depth = max(level for level, _ in rows) + 1
    data = []
    for level, name in rows:
        row = [''] * (level + 1)
        row[level] = name
        data.append(row)

    df = pd.DataFrame(data, columns=[f"Level {i+1}" for i in range(max_depth)])

    # Sauvegarder dans un buffer en mémoire
    with io.BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False)
        buffer.seek(0)
        with open(output_file, "wb") as f:
            f.write(buffer.read())
    print(f"Hiérarchie sauvegardée dans le fichier : {output_file}")

def replace_duplicates_with_nan(hierarchy_df):

    for column in hierarchy_df.columns:
        seen_values = []  # Liste pour suivre les valeurs uniques
        hierarchy_df[column] = hierarchy_df[column].apply(
            lambda x: x if x not in seen_values and not seen_values.append(x) else float('nan')
        )
    return hierarchy_df


def remove_duplicate_xlsx_files(hierarchy_df: pd.DataFrame) -> pd.DataFrame:
    """
    Supprime les doublons de fichiers Excel (.xlsx) dans un DataFrame hiérarchique.
    """
    seen_files = set()
    rows_to_keep = []

    for index, row in hierarchy_df.iterrows():
        file_name = row.iloc[-1]  # Supposons que le dernier niveau contient les noms de fichiers
        if isinstance(file_name, str) and file_name.endswith(".xlsx"):
            if file_name not in seen_files:
                seen_files.add(file_name)
                rows_to_keep.append(index)
        else:
            rows_to_keep.append(index)

    return hierarchy_df.loc[rows_to_keep].reset_index(drop=True)

def extract_hierarchy_from_zip(zip_buffer):
    """
    Extrait la hiérarchie des fichiers d'un ZIP en mémoire et structure la sortie en niveaux,
    avec suppression des doublons pour chaque niveau, sauf pour le Level 1.
    :param zip_buffer: Le buffer ZIP en mémoire.
    :return: Un DataFrame représentant la hiérarchie des fichiers dans le ZIP.
    """
    with zipfile.ZipFile(zip_buffer, 'r') as zipf:
        file_list = zipf.namelist()  # Liste des fichiers dans le ZIP

    # Construire la hiérarchie
    hierarchy = {}
    for file_path in file_list:
        parts = file_path.split('/')
        current_level = hierarchy
        for part in parts:
            if part not in current_level:
                current_level[part] = {}
            current_level = current_level[part]

    # Fonction récursive pour transformer la hiérarchie en lignes
    def traverse_hierarchy(node, depth=0, path=[]):
        rows = []
        for key, value in node.items():
            new_path = path + [''] * (depth - len(path)) + [key]
            rows.append(new_path)
            if isinstance(value, dict):  # Si c'est un dossier, continuer la traversée
                rows.extend(traverse_hierarchy(value, depth + 1, new_path))
            else:
                rows.append(new_path + [value])  # Ajouter les fichiers
        return rows

    # Extraire les lignes structurées
    rows = traverse_hierarchy(hierarchy)

    # Supprimer les doublons dans les colonnes en insérant des cellules vides pour éviter la répétition
    def remove_redundancy(rows):
        """
        Supprime les redondances dans les lignes du tableau en conservant uniquement
        les colonnes pertinentes.

        :param rows: Liste de listes représentant les lignes du tableau.
        :return: Liste de lignes nettoyées.
        """
        if not rows:
            raise ValueError("Le tableau 'rows' est vide. Vérifiez les données entrantes.")

        if not isinstance(rows[0], list):
            raise ValueError("Le tableau 'rows' doit être une liste de listes.")

        for col in range(1, len(rows[0])):  # Parcours des colonnes, sauf Level 1 (colonne 0)
            # Ajoutez votre logique pour traiter les colonnes ici.
            pass

        return rows

    if rows :
        rows = remove_redundancy(rows)

        # Trouver la profondeur maximale
        max_depth = max(len(row) for row in rows)

        # Compléter les lignes avec des colonnes vides jusqu'à la profondeur maximale
        structured_rows = [row + [''] * (max_depth - len(row)) for row in rows]

        # Construire un DataFrame
        
        df = pd.DataFrame(structured_rows, columns=[f"Level {i}" for i in range(0,max_depth)])
        return df

def process_generic(data, ref_paths, run_timestamp, export_type, zip_buffer, entity=None, currency=None):
    """
    Exemple générique d'une fonction de traitement écrivant dans le ZIP.
    """
    base_folder = f"RUN_{run_timestamp}_{export_type}"

    with zipfile.ZipFile(zip_buffer, 'a') as zipf:
        try:
            # Simulez une transformation et écrivez les résultats
            result_data = pd.DataFrame({"Col1": [1, 2], "Col2": [3, 4]})
            folder_path = f"{base_folder}/Example_Process"
            file_name = f"{folder_path}/Result.xlsx"

            # Créez un fichier temporaire et ajoutez-le au ZIP
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                result_data.to_excel(temp_file.name, index=False, engine="xlsxwriter")
                zipf.write(temp_file.name, arcname=file_name)

            print(f"Fichier ajouté au ZIP : {file_name}")
        except Exception as e:
            raise ValueError(f"Erreur lors de l'ajout des résultats au ZIP : {e}")



    
def count_entity_occurrences_from_df(export_type: str, hierarchy_df: pd.DataFrame, 
                                     chosen_entities: list = None, chosen_indicator: str = "ALL") -> tuple:
    """
    Compte les occurrences des entités au niveau 3 dans la hiérarchie et retourne deux DataFrames :
    1. Un DataFrame contenant les entités regroupées et leurs occurrences additionnées.
    2. Un DataFrame avec les colonnes 'indicateur' et 'Nombre occurrences' pour les mots-clés spécifiques.
    
    :param export_type: Le type d'export (e.g., ALL, BILAN, CONSO, GRAN).
    :param hierarchy_df: DataFrame contenant la hiérarchie (niveaux 1 à N).
    :param chosen_entities: Liste des entités choisies pour GRAN.
    :param chosen_indicator: Indicateur choisi pour GRAN (ou "ALL").
    :return: Tuple contenant deux DataFrames :
             - DataFrame regroupé des entités.
             - DataFrame des mots-clés spécifiques.
    """
    # Nettoyer les espaces inutiles des colonnes
    hierarchy_df.columns = hierarchy_df.columns.str.strip()

    if export_type == "GRAN":
        # Validation des entrées
        if not chosen_entities:
            raise ValueError("Pour le type d'export GRAN, vous devez fournir une liste d'entités choisies.")
        
        # Créer un DataFrame pour les entités choisies, chaque entité ayant une occurrence de 1
        result_df = pd.DataFrame({
            'Entités': chosen_entities,
            'Nombre occurrences': [1] * len(chosen_entities)
        })

        # Définir le Nombre occurrences pour les indicateurs
        num_entities = len(chosen_entities)
        if chosen_indicator == "ALL":
            # Tous les indicateurs reçoivent le même Nombre occurrences
            indicators_df = pd.DataFrame({
                'indicateur': ['LCR', 'AER', 'NSFR', 'QIS', 'ALMM'],
                'Nombre occurrences': [num_entities] * 5
            })
        else:
            # Seul l'indicateur choisi reçoit le Nombre occurrences
            indicators_df = pd.DataFrame({
                'indicateur': [chosen_indicator],
                'Nombre occurrences': [num_entities]
            })
        
        return result_df, indicators_df

    # Si l'export_type n'est pas GRAN, appliquer la logique standard
    try:
        all_index = hierarchy_df[hierarchy_df['Level 1'] == f'ALL'].index[0]
        eur_index = hierarchy_df[hierarchy_df['Level 1'] == f'EUR'].index[0]
        filtered_df_1 = hierarchy_df.iloc[all_index:eur_index + 1]
    except IndexError:
        st.write(hierarchy_df)
        raise ValueError(f"Les valeurs 'ALL' et 'EUR' ne sont pas présentes dans 'Level 1'.")

    # Filtrage de 'Level 2' avec 'Reports_by_entity'
    try:
        reports_by_entity_index = filtered_df_1[filtered_df_1['Level 2'] == 'Reports_by_entity'].index[0]
        filtered_df_2 = filtered_df_1.iloc[reports_by_entity_index:]
    except IndexError:
        raise ValueError(f"Le niveau 'Reports_by_entity' n'est pas trouvé dans 'Level 2'.")

    # Liste des entités et des comptages
    entity_list = []
    count_list = []

    # Variables pour suivre l'entité et ses occurrences
    last_entity = None
    current_count = 0

    # Variables pour compter les mots-clés spécifiques
    lcr_count = 0
    aer_count = 0
    nsfr_count = 0
    qis_count = 0
    almm_count = 0

    # Parcourir les lignes filtrées
    for _, row in filtered_df_2.iterrows():
        entity = row['Level 3']
        file_name = row['Level 4']

        # Si une nouvelle entité est rencontrée (valeur non nulle dans Level 3)
        if pd.notna(entity):
            # Sauvegarder le comptage de l'entité précédente si elle existe
            if last_entity is not None:
                entity_list.append(last_entity)
                count_list.append(current_count)
            # Réinitialiser pour la nouvelle entité
            last_entity = entity
            current_count = 0  # Réinitialiser le compteur pour la nouvelle entité

        # Si un fichier est présent dans Level 4, il est associé à l'entité courante
        if pd.notna(file_name):
            current_count += 1

            # Compter les occurrences des mots-clés spécifiques
            lcr_count += file_name.count('LCR_')
            aer_count += file_name.count('AER_')
            nsfr_count += file_name.count('NSFR_')
            qis_count += file_name.count('QIS_')
            almm_count += file_name.count('ALMM_')

    # Ajouter la dernière entité et son comptage (si applicable)
    if last_entity is not None:
        entity_list.append(last_entity)
        count_list.append(current_count)

    # Créer un DataFrame pour les entités et leur Nombre occurrences
    result_df = pd.DataFrame({
        'Entités': entity_list,
        'Nombre occurrences': count_list
    })

    # Regrouper les entités ayant le même nom et additionner leurs occurrences
    grouped_result_df = result_df.groupby("Entités", as_index=False).agg({"Nombre occurrences": "sum"})

    # Créer un DataFrame pour les mots-clés spécifiques
    indicators_df = pd.DataFrame({
        'indicateur': ['LCR', 'AER', 'NSFR', 'QIS', 'ALMM'],
        'Nombre occurrences': [lcr_count, aer_count, nsfr_count, qis_count, almm_count]
    })

    return grouped_result_df, indicators_df

def save_to_excel(data: pd.DataFrame, template_path: str, output_path: str, zip_buffer: zipfile.ZipFile):
    """
    Sauvegarde les données dans un fichier Excel en utilisant un template et ajoute le fichier dans un ZIP.
    """
    workbook = load_workbook(template_path)
    first_sheet_name = workbook.sheetnames[0]
    first_sheet = workbook[first_sheet_name]

    # Nettoyage de la feuille existante
    for row in first_sheet.iter_rows():
        for cell in row:
            cell.value = None

    # Ajout des données
    for col_index, col_name in enumerate(data.columns, start=1):
        first_sheet.cell(row=1, column=col_index, value=col_name)  # En-têtes
        for row_index, value in enumerate(data[col_name], start=2):
            first_sheet.cell(row=row_index, column=col_index, value=value)

    # Sauvegarde dans un fichier temporaire
    temp_file = io.BytesIO()
    workbook.save(temp_file)
    temp_file.seek(0)

    # Ajout dans le ZIP
    zip_buffer.writestr(output_path, temp_file.getvalue())
    print(f"Fichier sauvegardé dans le ZIP : {output_path}")

def save_excel_with_structure(
    processed_data: dict,
    template_path: str,
    entity_list: list,
    run_timestamp: str,
    export_type: str,
    zip_buffer: zipfile.ZipFile,
    entity: str = None,
    currency: str = "ALL"
):

    base_folder = f"RUN_{run_timestamp}_{export_type}"

    if not processed_data:
        st.warning("Aucune donnée à sauvegarder dans le ZIP.")
        return

    for currency_key, data in processed_data.items():
        if not isinstance(data, pd.DataFrame) or data.empty:
            st.warning(f"Aucune donnée disponible pour la devise '{currency_key}'.")
            continue

        # Créer les chemins pour les fichiers globaux
        global_folder = f"{base_folder}/{currency_key}/Reports_all_entities"
        global_file = f"{global_folder}/LCR_{export_type}_{currency_key}_All_Entities.xlsx"

        # Sauvegarder uniquement le fichier global si export_type == 'ALL'
        if export_type == 'ALL':
            save_to_excel(data, template_path, global_file, zip_buffer)
            continue

        # Sinon, créer également les fichiers par entité
        entity_folder = f"{base_folder}/{currency_key}/Reports_by_entity"
        save_to_excel(data, template_path, global_file, zip_buffer)

        for specific_entity in entity_list:
            entity_data = data[data["Ref_Entite.entité"] == specific_entity]
            if not entity_data.empty:
                entity_file = f"{entity_folder}/{specific_entity}/LCR_{export_type}_{currency_key}_{specific_entity}.xlsx"
                save_to_excel(entity_data, template_path, entity_file, zip_buffer)

    st.success("Données sauvegardées avec succès dans le ZIP.")


def generate_import_files(uploaded_data, run_timestamp, zip_buffer, import_folder):
        """
        Génère les fichiers d'import BILAN et CONSO pour les devises ALL, EUR, et USD,
        et les ajoute dans un dossier compressé au sein du ZIP final.

        :param uploaded_data: DataFrame chargé depuis le fichier téléchargé.
        :param run_timestamp: Timestamp pour nommer le dossier d'import.
        :param zip_buffer: Buffer ZIP où les fichiers seront ajoutés.
        :param import_folder: Nom du dossier où placer les fichiers dans le ZIP.
        """
        # Filtrages
        bilan_data = uploaded_data[uploaded_data["D_T1"] == "INTER"]
        conso_data = uploaded_data[uploaded_data["D_T1"] != "INTER"]

        # Itération sur les devises
        for curr in ["ALL", "EUR", "USD"]:
            if curr == "ALL":
                bilan_filtered = bilan_data
                conso_filtered = conso_data
            else:
                bilan_filtered = bilan_data[bilan_data["D_CU"] == curr]
                conso_filtered = conso_data[conso_data["D_CU"] == curr]

            # Génération des fichiers
            bilan_file = f"{import_folder}/IMPORT_BILAN_{curr}.xlsx"
            conso_file = f"{import_folder}/IMPORT_CONSO_{curr}.xlsx"

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_bilan_file:
                bilan_filtered.to_excel(temp_bilan_file.name, index=False, engine="xlsxwriter")
                with zipfile.ZipFile(zip_buffer, "a") as zipf:
                    zipf.write(temp_bilan_file.name, arcname=bilan_file)

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_conso_file:
                conso_filtered.to_excel(temp_conso_file.name, index=False, engine="xlsxwriter")
                with zipfile.ZipFile(zip_buffer, "a") as zipf:
                    zipf.write(temp_conso_file.name, arcname=conso_file)

        # Sauvegarder le fichier importé brut
        imported_file = f"{import_folder}/IMPORT_SOURCE.xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_imported_file:
            uploaded_data.to_excel(temp_imported_file.name, index=False, engine="xlsxwriter")
            with zipfile.ZipFile(zip_buffer, "a") as zipf:
                zipf.write(temp_imported_file.name, arcname=imported_file)

        print(f"Fichiers d'import sauvegardés dans le dossier : {import_folder}")

if __name__ == "__main__":
    st.title("HIBISCUS Generator.")
    custom_css = """
        <style>
        /* Cacher le menu principal et le footer */
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}

        /* Cacher le badge Viewer (Fork & GitHub) */
        .viewerBadge_container__1QSob {display: none !important;}
        .viewerBadge_link__yUdr6 {display: none !important;}

        /* Supprimer tout lien GitHub */
        a[href*="github.com"] {display: none !important;}
        </style>
    """
    st.markdown(custom_css, unsafe_allow_html=True)
    run_timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    # CSS pour forcer les boutons à occuper toute la largeur de la barre latéral
    # Initialiser l'état de navigation
    if "menu_choice" not in st.session_state:
        st.session_state.menu_choice = "Main"  # Page par défaut
    # CSS pour forcer les boutons à occuper toute la largeur de la barre latérale
    st.markdown(
        """
        <style>
        .sidebar-buttons {
            display: flex;
            flex-direction: column;
            gap: 10px;
            width : 100%;
        }
        .stButton button {
            all: unset;
            display: block;
            width: 100%;  /* Force le bouton à occuper toute la largeur */
            color: white;
            border: 1px solid grey;
            padding: 10px;
            border-radius: 5px;
            font-size: 16px;
            cursor: pointer;
            text-align: center;
            box-sizing: border-box; /* Assure que le padding est inclus dans la largeur */
            display :flex;
            align-items : left;
            justify-content : left;
        }
        .sidebar-buttons button:hover {
            background-color: #105ea2;
        }
        .stButton button.active {
            background-color: #0d4d8c;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # Barre latérale avec des boutons stylisés
    st.sidebar.title("Menu")
    st.sidebar.markdown('<div class="sidebar-buttons">', unsafe_allow_html=True)

    if st.sidebar.button("Main", key="main_button"):
        st.session_state.menu_choice = "Main"
    if st.sidebar.button("Export", key="export_button"):
        st.session_state.menu_choice = "Export"
    if st.sidebar.button("Fonctionnalités", key="features_button"):
        st.session_state.menu_choice = "Fonctionnalités"

    st.sidebar.markdown("</div>", unsafe_allow_html=True)




    if st.session_state.menu_choice == "Main" :
        # Introduction structurée pour l'application
        st.markdown(
            """
            <style>
                .intro-header {
                    font-size: 32px;
                    font-weight: bold;
                    color: #2E86C1;
                    text-align: center;
                }
                .intro-subheader {
                    font-size: 18px;
                    color: #5D6D7E;
                    text-align: center;
                    margin-bottom: 20px;
                }
                .intro-description {
                    font-size: 16px;
                    color: white;
                    line-height: 1.6;
                }
                .highlight {
                    color: #D35400;
                    font-weight: bold;
                }
            </style>
            """, unsafe_allow_html=True
        )

        st.markdown('<div class="intro-header">🌟 HIBISCUS Generator 🌟</div>', unsafe_allow_html=True)
        st.markdown('<div class="intro-subheader">Un outil avancé pour générer des rapports financiers dynamiques</div>', unsafe_allow_html=True)

        st.markdown(
            """
            <div class="intro-description">
                Bienvenue dans <span class="highlight">HIBISCUS Generator</span>, une application web conçue pour simplifier et 
                automatiser le traitement des données financières hiérarchiques. Grâce à cet outil, vous pouvez :
                <ul>
                    <li>📊 Générer des rapports structurés pour différents types d'exports (<span class="highlight">ALL</span>, <span class="highlight">BILAN</span>, <span class="highlight">CONSO</span>, et <span class="highlight">GRAN</span>).</li>
                    <li>⚙️ Appliquer des processus financiers avancés comme <span class="highlight">NSFR</span>, <span class="highlight">LCR</span>, <span class="highlight">QIS</span>, <span class="highlight">ALMM</span>, et <span class="highlight">AER</span>.</li>
                    <li>📦 Exporter les résultats sous forme de fichiers compressés directement téléchargeables.</li>
                </ul>
                Laissez-vous guider par notre interface intuitive pour réaliser vos analyses financières en toute simplicité.
            </div>
            """,
            unsafe_allow_html=True
        )

    if st.session_state.menu_choice == "Export":
        
        st.subheader('Export des données:')
        
        st.markdown("""
                <style>
                .feature-header {
                    font-size: 24px;
                    font-weight: bold;
                    color: #2E86C1;
                    margin-top: 20px;
                    text-align: center;
                }
                .feature-description {
                    font-size: 16px;
                    color: #34495E;
                    line-height: 1.6;
                    margin-bottom: 20px;
                }
                .bold{
                    font-weight: bold;
                    color : lightgrey;
                }
                </style>
            """, unsafe_allow_html=True)
                
        # Téléchargement du fichier
        uploaded_file = st.sidebar.file_uploader("Téléchargez votre fichier Excel hiérarchique", type=["xlsx"])
        if uploaded_file is not None:
            try:
                # Charger le fichier en mémoire
                df = pd.read_excel(uploaded_file)

                # Afficher les données pour confirmation
                st.write("Aperçu des données :", df.head())

                # Exécuter vos traitements ici
                st.success("Fichier chargé et traité avec succès !")

            except Exception as e:
                st.error(f"Erreur lors du chargement du fichier : {e}")
        else:
            st.info("Veuillez télécharger un fichier pour commencer.")
        export_type = st.sidebar.selectbox("Choisissez le type d'export :", ["ALL", "BILAN", "CONSO", "GRAN"])
        run_timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        
        # Paramètres pour GRAN
        entity, currency, indicator, selected_processes = None, None, None, "ALL"
        if export_type == "GRAN":
            # Indicateur, Entité et Devise pour le GRAN
            indicator = st.sidebar.selectbox("Choisissez la vue :", ["ALL", "BILAN", "CONSO"])
            entity = st.sidebar.selectbox("Choisissez l'entité spécifique :", ["ALL"] + Entity_List)
            currency = st.sidebar.selectbox("Devise spécifique :", ["ALL","EUR", "USD"])
            selected_processes = st.sidebar.multiselect(
                "Sélectionnez les processus à exécuter :",
                ["ALL", "NSFR", "LCR", "QIS", "ALMM", "AER"],
                default="ALL"
            )
        if st.sidebar.button("Lancer le traitement"):
            st.info('OUAIS !')
            if uploaded_file:
                uploaded_data = pd.read_excel(uploaded_file)
                missing_columns = [col for col in expected_columns if col not in uploaded_data.columns]
                if missing_columns:
                    st.error("Certaines colonnes attendues sont manquantes dans le fichier :")

                    # Affichage des colonnes manquantes dans un tableau
                    missing_df = pd.DataFrame(
                        {"Colonnes manquantes": missing_columns}
                    )
                    st.markdown(
                        """
                        <style>
                            .missing-table {
                                border-radius: 5px;
                                padding: 10px;
                                margin-top: 10px;
                                margin-bottom: 10px;
                                box-shadow: 2px 2px 5px rgba(0, 0, 0, 0.1);
                            }
                            .missing-table h3 {
                                color: lightgrey;
                                margin-bottom: 10px;
                            }
                        </style>
                        """,
                        unsafe_allow_html=True,
                    )

                    # Convertir le tableau en HTML et l'afficher
                    st.markdown(
                        f"""
                        <div class="missing-table">
                            <h3>Colonnes manquantes :</h3>
                            {missing_df.to_html(index=False, escape=False, justify="center")}
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                else:
                    try:
                        # Initialiser le buffer ZIP
                        zip_buffer = io.BytesIO()
                        
                        import_folder = f"import_{run_timestamp}"

                        with tempfile.TemporaryDirectory() as temp_dir:
                            # Sauvegarder le fichier téléchargé
                            input_file_path = os.path.join(temp_dir, "uploaded_hierarchy.xlsx")
                            with open(input_file_path, "wb") as f:
                                f.write(uploaded_file.getbuffer())

                            # Barre de progression et état actuel
                            progress_bar = st.progress(0)
                            current_task_placeholder = st.empty()

                            # Étape 1 : Prétraitement des données
                            current_task_placeholder.text("Prétraitement des données...")
                            generate_import_files(uploaded_data, run_timestamp, zip_buffer,import_folder)
                            preprocessed_data = preprocess_all_data(
                                data_path=input_file_path,
                                ref_entite_path="./Ref 2/ref_entite.xlsx",
                                ref_transfo_path="./Ref 2/ref_transfo_l1.xlsx",
                                ref_lcr_path="./Ref 2/ref_lcr.xlsx",
                                ref_adf_lcr_path="./Ref 2/ref_lcr_adf.xlsx",
                                input_excel_path="./Livrable/Templates/LCR_Template.xlsx",
                                run_timestamp=run_timestamp,
                                export_type=export_type,
                                currency=currency,
                            )
                            progress_bar.progress(20)
                            
                            # Vérification du type de données retournées
                            if export_type == "GRAN":
                                if "filtered_data" in preprocessed_data:
                                    gran_data = preprocessed_data["filtered_data"]
                                    generated_import_files = gran_data  # Le résultat contient les chemins des fichiers générés
                                    
                                else:
                                    st.write("Les données filtrées pour GRAN sont absentes.")
                            else:
                                # Pour les autres types d'export
                                generated_import_files = preprocessed_data  # Le résultat contient les chemins des fichiers générés
                                
                            progress_bar.progress(40)

                            # Étape 2 : Exécution des processus
                            current_task_placeholder.text("Exécution des processus...")
                            processes = {
                                "NSFR": {
                                    "func": process_nsfr,
                                    "args": [
                                        preprocessed_data, input_file_path, "./Ref 2/ref_entite.xlsx",
                                        "./Ref 2/ref_transfo_l1.xlsx", "./Ref 2/ref_nsfr.xlsx",
                                        "./Ref 2/ref_nsfr_adf.xlsx", "./Ref 2/ref_dzone_nsfr.xlsx",
                                        "./Livrable/Templates/NSFR_Template.xlsx", run_timestamp,
                                        export_type, zip_buffer, entity, currency, indicator
                                    ],
                                },
                                "LCR": {
                                    "func": process_lcr,
                                    "args": [
                                        preprocessed_data, input_file_path, "./Ref 2/ref_entite.xlsx",
                                        "./Ref 2/ref_transfo_l1.xlsx", "./Ref 2/ref_lcr.xlsx",
                                        "./Ref 2/ref_lcr_adf.xlsx", "./Livrable/Templates/LCR_Template.xlsx",
                                        run_timestamp, export_type, zip_buffer, entity, currency, indicator
                                    ],
                                },
                                "QIS": {
                                    "func": process_qis,
                                    "args": [
                                        preprocessed_data, input_file_path, "./Ref 2/ref_entite.xlsx",
                                        "./Ref 2/ref_transfo_l1.xlsx", "./Ref 2/Ref_QIS.xlsx",
                                        "./Ref 2/ref_nsfr_adf.xlsx", "./Ref 2/ref_dzone_nsfr.xlsx",
                                        "./Livrable/Templates/QIS_Template.xlsx", run_timestamp,
                                        export_type, zip_buffer, entity, currency, indicator
                                    ],
                                },
                                "ALMM": {
                                    "func": process_almm,
                                    "args": [
                                        preprocessed_data, input_file_path, "./Ref 2/ref_entite.xlsx",
                                        "./Ref 2/ref_transfo_l1.xlsx", "./Ref 2/ref_nsfr.xlsx",
                                        "./Ref 2/ref_nsfr_adf.xlsx", "./Ref 2/ref_dzone_nsfr.xlsx",
                                        "./Livrable/Templates/ALMM_Template.xlsx", run_timestamp,
                                        export_type, zip_buffer, entity, currency, indicator
                                    ],
                                },
                                "AER": {
                                    "func": process_aer,
                                    "args": [
                                        preprocessed_data, input_file_path, "./Ref 2/ref_entite.xlsx",
                                        "./Ref 2/ref_transfo_l1.xlsx", "./Ref 2/ref_aer.xlsx",
                                        "./Ref 2/ref_aer_adf.xlsx", "./Livrable/Templates/AER_Template.xlsx",
                                        run_timestamp, export_type, zip_buffer, entity, currency, indicator
                                    ],
                                },
                            }

                            if selected_processes == "ALL":
                                selected_processes = list(processes.keys())

                            step_progress = 40
                            for i, process_name in enumerate(selected_processes, start=1):
                                current_task_placeholder.text(f"Exécution du processus {process_name}...")
                                process_info = processes.get(process_name)
                                if process_info:
                                    process_info["func"](*process_info["args"])
                                else:
                                    print(f"Processus '{process_name}' non reconnu.")
                                progress_bar.progress(step_progress + (i * int(30 / len(selected_processes))))

                            current_task_placeholder.text("Génération des fichiers de hiérarchie...")
                            hierarchy_file_path = os.path.join(temp_dir, "hierarchy_all.xlsx")
                            hierarchy_df = extract_hierarchy_from_zip(zip_buffer)
                            hierarchy_df = replace_duplicates_with_nan(hierarchy_df)

                            hierarchy_df.to_excel(hierarchy_file_path, index=False)

                            current_task_placeholder.text("Ajout des fichiers au ZIP final...")
                            with zipfile.ZipFile(zip_buffer, "a") as zipf:

                                # Ajouter le fichier de hiérarchie
                                zipf.write(hierarchy_file_path, arcname="hierarchy_all.xlsx")
                                
                                # Ajouter le fichier des occurrences uniquement si ce n'est pas GRAN
                                if export_type == 'GRAN':
                                    if entity == "ALL":
                                        chosen_entities = Entity_List
                                    else:
                                        chosen_entities = [entity]

                                    if "ALL" in selected_processes:
                                        chosen_indicator = "ALL"
                                    else:
                                        # Concaténer les processus sélectionnés pour l'indicateur
                                        chosen_indicator = ", ".join(selected_processes)

                                    # Appel de la fonction pour GRAN
                                    grouped_count_df, indicators_df = count_entity_occurrences_from_df(
                                        export_type="GRAN",
                                        hierarchy_df=hierarchy_df,
                                        chosen_entities=chosen_entities,
                                        chosen_indicator=chosen_indicator
                                    )

                                    # Ajouter les entités manquantes avec 0 occurrences au DataFrame des entités
                                    all_entities = set(Entity_List)
                                    existing_entities = set(grouped_count_df["Entités"])
                                    missing_entities = all_entities - existing_entities

                                    # Ajouter les entités manquantes au DataFrame
                                    missing_df = pd.DataFrame({
                                        "Entités": list(missing_entities),
                                        "Nombre occurrences": [0] * len(missing_entities)
                                    })
                                    grouped_count_df = pd.concat([grouped_count_df, missing_df], ignore_index=True)

                                    # Générer le fichier Excel avec les résultats
                                    count_file_path = os.path.join(temp_dir, "count_gran.xlsx")
                                    with pd.ExcelWriter(count_file_path, engine='openpyxl') as writer:
                                        # Écrire le DataFrame des entités
                                        grouped_count_df.to_excel(writer, index=False, sheet_name="Résultats", startrow=0)

                                        # Ajouter 5 lignes vides avant le DataFrame des indicateurs
                                        start_row = len(grouped_count_df) + 6  # 1 ligne pour l'en-tête + 5 lignes vides
                                        indicators_df.to_excel(writer, index=False, sheet_name="Résultats", startrow=start_row)

                                    # Ajouter le fichier Excel dans le ZIP
                                    zipf.write(count_file_path, arcname="KPI_GRAN.xlsx")
                                
                                if export_type != "GRAN" and export_type != "ALL":
                                    count_file_path = os.path.join(temp_dir, "count_all.xlsx")
                                                                            
                                    grouped_count_df, indicators_df = count_entity_occurrences_from_df(export_type, hierarchy_df)
                                    
                                    # Ajouter les entités manquantes avec 0 occurrences au DataFrame des entités
                                    all_entities = set(Entity_List)
                                    existing_entities = set(grouped_count_df["Entités"])
                                    missing_entities = all_entities - existing_entities

                                    # Ajouter les entités manquantes au DataFrame
                                    missing_df = pd.DataFrame({
                                        "Entités": list(missing_entities),
                                        "Nombre occurrences": [0] * len(missing_entities)
                                    })
                                    grouped_count_df = pd.concat([grouped_count_df, missing_df], ignore_index=True)

                                    # Écrire les deux DataFrames dans un fichier Excel
                                    with pd.ExcelWriter(count_file_path, engine='openpyxl') as writer:
                                        # Écrire le premier DataFrame
                                        grouped_count_df.to_excel(writer, index=False, sheet_name="Résultats", startrow=0)
                                        
                                        # Ajouter 5 lignes vides avant le second DataFrame
                                        start_row = len(grouped_count_df) + 6  # 1 ligne pour l'en-tête + 5 lignes vides
                                        indicators_df.to_excel(writer, index=False, sheet_name="Résultats", startrow=start_row)
                                    
                                    # Ajouter le fichier Excel dans le ZIP
                                    zipf.write(count_file_path, arcname="KPI.xlsx")

                            progress_bar.progress(90)

                            # Proposer le téléchargement
                            zip_buffer.seek(0)
                            st.download_button(
                                label="Télécharger les résultats (ZIP)",
                                data=zip_buffer.getvalue(),
                                file_name=f"RUN_{run_timestamp}_{export_type}.zip",
                                mime="application/zip",
                            )
                            progress_bar.progress(100)
                            current_task_placeholder.success("Traitement terminé avec succès !")

                    except Exception as e:
                        import traceback
                        current_task_placeholder.text(f"Une erreur est survenue : {e}")
                        st.text("Traceback détaillé :")
                        st.text(traceback.format_exc())

        else:
            st.markdown('<div class="feature-description bold">Importez un fichier et choisissez la méthode pour exporter et autres filtres si nécessaire.</div>', unsafe_allow_html=True)
            st.info('NOP !') 


    elif st.session_state.menu_choice == "Fonctionnalités":
        st.subheader("Fonctionnalités de l'application")

        # Style global pour centrer le contenu des colonnes
        st.markdown(
            """
            <style>
            .custom-column2 {
                text-align: center;
                padding: 10px;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )

        # Titre principal
        st.subheader("📄 Téléchargement de Documents")
        
        

        # Conteneur pour l'alignement des colonnes
        with st.container():
            # Colonnes
            doc_col1, doc_col2, doc_col3 = st.columns(3)

            # Bouton 1 : PowerPoint
            with doc_col1:
                st.markdown(
                    """
                    <div class="custom-column2">
                        <strong>Présentation du Projet (PowerPoint)</strong>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
                try:
                    with open("./Livrable/Tool_System/Hibiscus_livrable_version_client.pptx", "rb") as file:
                        st.download_button(
                            label="📤 Télécharger PowerPoint",
                            data=file,
                            file_name="presentation_hibiscus.pptx",
                            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                            key="download_powerpoint",
                        )
                except FileNotFoundError:
                    st.error("Le fichier PowerPoint est introuvable.")
                except Exception as e:
                    st.error(f"Erreur lors du téléchargement : {e}")

            # Bouton 2 : PDF Logique
            with doc_col2:
                st.markdown(
                    """
                    <div class="custom-column2">
                        <strong>Présentation du guide utilisateur (PDF)</strong>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
                try:
                    with open("./Livrable/Tool_System/User Guide.pdf", "rb") as file:
                        st.download_button(
                            label="📤 Télécharger User Guide ",
                            data=file,
                            file_name="logic_documentation.pdf",
                            mime="application/pdf",
                            key="download_pdf_logic",
                        )
                except FileNotFoundError:
                    st.error("Le fichier PDF est introuvable.")
                except Exception as e:
                    st.error(f"Erreur lors du téléchargement : {e}")

            # Bouton 3 : Nouveau PDF
            with doc_col3:
                st.markdown(
                    """
                    <div class="custom-column2">
                        <strong>Documentation Technique (PDF)</strong>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
                try:
                    with open("./Livrable/Tool_System/Rapport_Specification_Fonctionnelle_Hibiscus.pdf", "rb") as file:
                        st.download_button(
                            label="📤 Télécharger PDF Technique",
                            data=file,
                            file_name="technical_documentation.pdf",
                            mime="application/pdf",
                            key="download_pdf_technique",
                        )
                except FileNotFoundError:
                    st.error("Le fichier PDF technique est introuvable.")
                except Exception as e:
                    st.error(f"Erreur lors du téléchargement : {e}")
